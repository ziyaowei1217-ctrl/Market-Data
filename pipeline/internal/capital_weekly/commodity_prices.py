from __future__ import annotations

import calendar
import io
import json
import math
import re
import unicodedata
from collections.abc import Mapping
from datetime import date, datetime

from openpyxl import load_workbook


WORLD_BANK_MONTHLY_PRICE_HEADINGS = frozenset({
    "monthly prices",
    "monthly prices in nominal us dollars, 1960 to present",
})


def _normalized_label(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    return " ".join(text.replace("\xa0", " ").split()).casefold()


def parse_eia_price_series(
    text: str,
    series_code: str,
    expected_unit: str,
) -> list[dict]:
    payload = json.loads(text)
    data = payload.get("response", {}).get("data")
    if not isinstance(data, list):
        raise ValueError("EIA response has no data array")

    rows = []
    seen_dates: set[date] = set()
    for raw in data:
        if not isinstance(raw, dict) or raw.get("series") != series_code:
            continue
        raw_unit = raw.get("unit") or raw.get("units")
        unit = str(raw_unit or "").strip()
        if unit != expected_unit:
            raise ValueError(
                f"Unexpected EIA unit for {series_code}: {unit!r}; "
                f"expected {expected_unit!r}"
            )
        try:
            observation_date = date.fromisoformat(str(raw["period"]).strip()[:10])
            value = float(raw["value"])
        except (KeyError, TypeError, ValueError) as error:
            raise ValueError(f"Invalid EIA observation for {series_code}") from error
        if observation_date in seen_dates:
            raise ValueError(
                f"Duplicate EIA date for {series_code}: {observation_date.isoformat()}"
            )
        if not math.isfinite(value):
            raise ValueError(
                f"Invalid EIA value for {series_code} on {observation_date.isoformat()}"
            )
        seen_dates.add(observation_date)
        rows.append({"date": observation_date, "value": value, "unit": unit})

    if not rows:
        raise ValueError(f"EIA response has no observations for series {series_code}")
    return sorted(rows, key=lambda row: row["date"])


def _month_end(value: object) -> date | None:
    if isinstance(value, datetime):
        year, month = value.year, value.month
    elif isinstance(value, date):
        year, month = value.year, value.month
    else:
        text = str(value or "").strip()
        match = re.fullmatch(r"(\d{4})\s*[Mm-]\s*(\d{1,2})", text)
        if match:
            year, month = map(int, match.groups())
        else:
            try:
                parsed = datetime.fromisoformat(text)
            except ValueError:
                return None
            year, month = parsed.year, parsed.month
    try:
        return date(year, month, calendar.monthrange(year, month)[1])
    except ValueError:
        return None


def parse_world_bank_monthly_prices(
    content: bytes,
    columns: Mapping[str, str],
) -> dict[str, list[dict]]:
    if not isinstance(columns, Mapping) or not columns:
        raise ValueError("World Bank columns must map labels to expected units")
    requested = {
        _normalized_label(label): (label, expected_unit)
        for label, expected_unit in columns.items()
    }
    if len(requested) != len(columns):
        raise ValueError("World Bank requested columns are not unique after normalization")

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        candidate_sheets = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if any(
                _normalized_label(cell) in WORLD_BANK_MONTHLY_PRICE_HEADINGS
                for row in rows
                for cell in row
            ):
                candidate_sheets.append((sheet.title, rows))
        if len(candidate_sheets) != 1:
            raise ValueError("World Bank workbook must contain one Monthly Prices heading")
        _, rows = candidate_sheets[0]

        header_index = None
        column_indexes: dict[str, int] = {}
        for row_index, row in enumerate(rows):
            normalized = [_normalized_label(cell) for cell in row]
            positions = {}
            for label, (source_label, _) in requested.items():
                matches = [
                    index
                    for index, value in enumerate(normalized)
                    if value == label
                ]
                if len(matches) > 1:
                    raise ValueError(
                        "World Bank workbook has duplicate requested column: "
                        f"{source_label}"
                    )
                if matches:
                    positions[label] = matches[0]
            explicit_date_header = any(
                value in {"date", "month"} for value in normalized
            )
            current_blank_date_header = (
                bool(positions)
                and bool(normalized)
                and not normalized[0]
                and row_index + 2 < len(rows)
                and _month_end(rows[row_index + 2][0]) is not None
            )
            if explicit_date_header or current_blank_date_header:
                header_index = row_index
                column_indexes = positions
                break
        if header_index is None:
            raise ValueError("World Bank workbook is missing the monthly price header")
        missing = [
            requested[label][0]
            for label in requested
            if label not in column_indexes
        ]
        if missing:
            raise ValueError(
                "World Bank workbook is missing requested column(s): "
                + ", ".join(missing)
            )

        header = rows[header_index]
        date_index = next(
            (
                index
                for index, value in enumerate(header)
                if _normalized_label(value) in {"date", "month"}
            ),
            0,
        )
        if header_index + 1 >= len(rows):
            raise ValueError("World Bank workbook is missing the unit row")
        unit_row = rows[header_index + 1]
        native_units: dict[str, str] = {}
        for normalized, column_index in column_indexes.items():
            label, expected_unit = requested[normalized]
            unit = str(unit_row[column_index] or "").strip()
            if re.fullmatch(r"\([^()]+\)", unit):
                unit = unit[1:-1].strip()
            if unit != expected_unit:
                raise ValueError(
                    f"Unexpected World Bank unit for {label}: {unit!r}; "
                    f"expected {expected_unit!r}"
                )
            native_units[label] = unit

        parsed = {label: [] for label in columns}
        seen_dates = {label: set() for label in columns}
        for row in rows[header_index + 2 :]:
            if date_index >= len(row):
                continue
            observation_date = _month_end(row[date_index])
            if observation_date is None:
                continue
            for normalized, column_index in column_indexes.items():
                label, _ = requested[normalized]
                if column_index >= len(row):
                    raise ValueError(
                        f"Invalid World Bank value for {label} on "
                        f"{observation_date.isoformat()}"
                    )
                raw_value = row[column_index]
                if isinstance(raw_value, str) and raw_value.strip() == "…":
                    continue
                try:
                    value = float(raw_value)
                except (TypeError, ValueError) as error:
                    raise ValueError(
                        f"Invalid World Bank value for {label} on "
                        f"{observation_date.isoformat()}"
                    ) from error
                if not math.isfinite(value):
                    raise ValueError(
                        f"Invalid World Bank value for {label} on "
                        f"{observation_date.isoformat()}"
                    )
                if observation_date in seen_dates[label]:
                    raise ValueError(
                        f"Duplicate World Bank date for {label}: "
                        f"{observation_date.isoformat()}"
                    )
                seen_dates[label].add(observation_date)
                parsed[label].append(
                    {
                        "date": observation_date,
                        "value": value,
                        "unit": native_units[label],
                    }
                )
        for label, observations in parsed.items():
            if not observations:
                raise ValueError(
                    f"World Bank workbook has no finite observations for {label}"
                )
            observations.sort(key=lambda row: row["date"])
        return parsed
    finally:
        workbook.close()


__all__ = [
    "parse_eia_price_series",
    "parse_world_bank_monthly_prices",
]
