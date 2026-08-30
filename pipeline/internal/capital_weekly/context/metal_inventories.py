from __future__ import annotations

from datetime import datetime
import hashlib
from io import BytesIO
import json
import math
import re
from typing import Any, Mapping

from openpyxl import load_workbook
from pypdf import PdfReader
import xlrd


OOXML_MAGIC = b"PK\x03\x04"
OLE2_MAGIC = bytes.fromhex("d0cf11e0a1b11ae1")
COMEX_COLUMNS = (
    "PREV TOTAL",
    "RECEIVED",
    "WITHDRAWN",
    "NET CHANGE",
    "ADJUSTMENT",
    "TOTAL TODAY",
)


def _required(spec: Mapping[str, Any], key: str) -> str:
    value = str(spec.get(key) or "").strip()
    if not value:
        raise ValueError(f"COMEX spec missing {key}")
    return value


def _container_name(content: bytes) -> str:
    if content.startswith(OOXML_MAGIC):
        return "ooxml-xlsx"
    if content.startswith(OLE2_MAGIC):
        return "ole2-biff8"
    raise ValueError(
        "Unsupported COMEX workbook container; expected OOXML or OLE2/BIFF8"
    )


def _workbook_rows(content: bytes, expected_sheet: str) -> tuple[str, list[list[Any]]]:
    container = _container_name(content)
    if container == "ooxml-xlsx":
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if workbook.sheetnames != [expected_sheet]:
            raise ValueError(
                f"COMEX sheet mismatch: expected only {expected_sheet!r}, "
                f"got {workbook.sheetnames!r}"
            )
        sheet = workbook[expected_sheet]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return container, rows

    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    if workbook.sheet_names() != [expected_sheet]:
        raise ValueError(
            f"COMEX sheet mismatch: expected only {expected_sheet!r}, "
            f"got {workbook.sheet_names()!r}"
        )
    sheet = workbook.sheet_by_name(expected_sheet)
    rows = [sheet.row_values(index) for index in range(sheet.nrows)]
    workbook.release_resources()
    return container, rows


def _text(value: Any) -> str:
    return str(value or "").strip()


def _find_unique_first_cell(rows: list[list[Any]], expected: str, label: str) -> int:
    matches = [index for index, row in enumerate(rows) if row and _text(row[0]) == expected]
    if len(matches) != 1:
        raise ValueError(
            f"COMEX {label} mismatch: expected exactly one {expected!r}, "
            f"found {len(matches)}"
        )
    return matches[0]


def _finite_number(value: Any, context: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"COMEX {context} must be a finite numeric value")
    result = float(value)
    if not math.isfinite(result):
        raise ValueError(f"COMEX {context} must be a finite numeric value")
    return result


def _reconciles(registered: float, eligible: float, total: float) -> bool:
    return math.isclose(registered + eligible, total, rel_tol=0.0, abs_tol=1e-6)


def comex_schema_signature(content: bytes, spec: Mapping[str, Any]) -> str:
    container, _rows = _workbook_rows(content, _required(spec, "expected_sheet"))
    schema = {
        "sheet": _required(spec, "expected_sheet"),
        "commodity_title": _required(spec, "commodity_title"),
        "unit": _required(spec, "expected_unit"),
        "headers": [_required(spec, "location_header"), *COMEX_COLUMNS],
        "registered_total_label": _required(spec, "registered_total_label"),
        "eligible_total_label": _required(spec, "eligible_total_label"),
        "combined_total_label": _required(spec, "combined_total_label"),
    }
    digest = hashlib.sha256(
        json.dumps(schema, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    return f"{container}:sha256:{digest}"


def parse_comex_stocks(
    content: bytes,
    spec: Mapping[str, Any],
) -> list[dict[str, Any]]:
    expected_sheet = _required(spec, "expected_sheet")
    _container, rows = _workbook_rows(content, expected_sheet)
    title = _required(spec, "commodity_title")
    title_index = _find_unique_first_cell(rows, title, "commodity title")
    if title_index + 1 >= len(rows):
        raise ValueError("COMEX unit row is missing")
    unit = _text(rows[title_index + 1][0])
    expected_unit = _required(spec, "expected_unit")
    if unit != expected_unit:
        raise ValueError(
            f"COMEX unit mismatch: expected {expected_unit!r}, got {unit!r}"
        )
    report_cell = _text(rows[title_index][6] if len(rows[title_index]) > 6 else "")
    report_match = re.fullmatch(r"Report Date: (\d{1,2}/\d{1,2}/\d{4})", report_cell)
    if not report_match:
        raise ValueError(f"COMEX report date mismatch: {report_cell!r}")
    report_date = datetime.strptime(report_match.group(1), "%m/%d/%Y").date()

    location_header = _required(spec, "location_header")
    header_index = _find_unique_first_cell(rows, location_header, "header")
    actual_headers = tuple(
        _text(rows[header_index][index] if len(rows[header_index]) > index else "")
        for index in (2, 3, 4, 5, 6, 7)
    )
    if actual_headers != COMEX_COLUMNS:
        raise ValueError(
            f"COMEX header mismatch: expected {COMEX_COLUMNS!r}, got {actual_headers!r}"
        )

    grand_labels = {
        _required(spec, "registered_total_label"): "registered",
        _required(spec, "eligible_total_label"): "eligible",
        _required(spec, "combined_total_label"): "total",
    }
    local_labels = {
        "registered": "registered",
        "registered (warranted)": "registered",
        "eligible": "eligible",
        "eligible (non-warranted)": "eligible",
        "total": "total",
    }
    parsed: list[dict[str, Any]] = []
    current_location: str | None = None
    grand_seen: set[str] = set()
    for row_number, row in enumerate(rows[header_index + 1 :], header_index + 2):
        label = _text(row[0] if row else "")
        if not label:
            continue
        if label in grand_labels:
            inventory_type = grand_labels[label]
            value = _finite_number(
                row[7] if len(row) > 7 else None,
                f"row {row_number} {label}",
            )
            parsed.append(
                {
                    "commodity_code": _required(spec, "commodity_code"),
                    "commodity_family": _required(spec, "commodity_family"),
                    "report_date": report_date,
                    "unit": unit,
                    "scope": "exchange",
                    "location": None,
                    "inventory_type": inventory_type,
                    "value": value,
                }
            )
            grand_seen.add(inventory_type)
            if inventory_type == "total":
                break
            continue
        normalized = label.casefold()
        if normalized == "pledged":
            continue
        if normalized in local_labels:
            if current_location is None:
                raise ValueError(f"COMEX inventory row {row_number} has no location")
            inventory_type = local_labels[normalized]
            parsed.append(
                {
                    "commodity_code": _required(spec, "commodity_code"),
                    "commodity_family": _required(spec, "commodity_family"),
                    "report_date": report_date,
                    "unit": unit,
                    "scope": "location",
                    "location": current_location,
                    "inventory_type": inventory_type,
                    "value": _finite_number(
                        row[7] if len(row) > 7 else None,
                        f"row {row_number} {label}",
                    ),
                }
            )
            continue
        current_location = label

    if grand_seen != {"registered", "eligible", "total"}:
        raise ValueError("COMEX workbook is missing one or more exchange totals")
    location_rows = [row for row in parsed if row["scope"] == "location"]
    locations = {str(row["location"]) for row in location_rows}
    for location in sorted(locations):
        values = {
            row["inventory_type"]: row["value"]
            for row in location_rows
            if row["location"] == location
        }
        if set(values) != {"registered", "eligible", "total"}:
            raise ValueError(f"COMEX {location} inventory rows are incomplete")
        if not _reconciles(values["registered"], values["eligible"], values["total"]):
            raise ValueError(f"COMEX {location} registered + eligible does not reconcile")

    exchange = {
        row["inventory_type"]: row["value"]
        for row in parsed
        if row["scope"] == "exchange"
    }
    if not _reconciles(exchange["registered"], exchange["eligible"], exchange["total"]):
        raise ValueError("COMEX exchange registered + eligible does not reconcile")
    for inventory_type in ("registered", "eligible", "total"):
        subtotal = sum(
            row["value"]
            for row in location_rows
            if row["inventory_type"] == inventory_type
        )
        if not math.isclose(
            subtotal,
            exchange[inventory_type],
            rel_tol=0.0,
            abs_tol=1e-6,
        ):
            raise ValueError(
                f"COMEX location {inventory_type} rows do not reconcile to exchange total"
            )
    return parsed


def parse_usgs_mcs_text(
    text: str,
    spec: Mapping[str, Any],
) -> list[dict[str, Any]]:
    commodity_title = str(spec.get("commodity_title") or "").strip()
    expected_unit = str(spec.get("expected_unit") or "").strip()
    publication_month = str(spec.get("publication_month") or "").strip()
    if not commodity_title or not expected_unit or not publication_month:
        raise ValueError("USGS spec is incomplete")
    title_pattern = rf"(?m)^\s*{re.escape(commodity_title)}\s*$"
    if not re.search(title_pattern, text):
        raise ValueError(f"USGS commodity title mismatch: {commodity_title!r}")
    unit_match = re.search(
        r"\(Data in ([^)]+), unless otherwise specified\)",
        text,
    )
    actual_unit = unit_match.group(1).strip() if unit_match else ""
    actual_unit = re.sub(
        r",(?:\d+|[⁰¹²³⁴⁵⁶⁷⁸⁹]+)\s+",
        ", ",
        actual_unit,
    )
    if actual_unit != expected_unit:
        raise ValueError(
            f"USGS unit mismatch: expected {expected_unit!r}, got {actual_unit!r}"
        )
    if f"Mineral Commodity Summaries, {publication_month}" not in text:
        raise ValueError("USGS publication month mismatch")
    table_kind = str(spec.get("table_kind") or "").strip()
    if table_kind == "mine_refinery_reserves":
        section = "World Mine and Refinery Production and Reserves:"
        expected_count = 5
        production_index = 1
    elif table_kind == "mine_reserves":
        section = "World Mine Production and Reserves:"
        expected_count = 3
        production_index = 1
    else:
        raise ValueError(f"Unsupported USGS table kind: {table_kind or 'blank'}")
    if section not in text:
        raise ValueError(f"USGS table heading mismatch: {section}")
    total_match = re.search(r"(?m)^\s*World total \(rounded\)\s+([^\r\n]+)$", text)
    number_tokens = (
        re.findall(r"(?:\d{1,3}(?:,\d{3})+|\d+)", total_match.group(1))
        if total_match
        else []
    )
    if len(number_tokens) != expected_count:
        raise ValueError(
            f"USGS world total has {len(number_tokens)} values; expected {expected_count}"
        )
    values = [float(token.replace(",", "")) for token in number_tokens]
    if not all(math.isfinite(value) for value in values):
        raise ValueError("USGS world total values must be finite")
    common = {
        "commodity_code": str(spec.get("commodity_code") or "").strip(),
        "commodity_family": str(spec.get("commodity_family") or "").strip(),
        "unit": expected_unit,
        "reference_period": str(spec.get("reference_year") or "").strip(),
    }
    return [
        {**common, "measurement": "mine_production", "value": values[production_index]},
        {**common, "measurement": "reserves", "value": values[-1]},
    ]


def parse_usgs_mcs_pdf(
    content: bytes,
    spec: Mapping[str, Any],
) -> list[dict[str, Any]]:
    if not content.startswith(b"%PDF-"):
        raise ValueError("USGS artifact is not a PDF")
    reader = PdfReader(BytesIO(content), strict=True)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return parse_usgs_mcs_text(text, spec)


__all__ = [
    "comex_schema_signature",
    "parse_comex_stocks",
    "parse_usgs_mcs_pdf",
    "parse_usgs_mcs_text",
]
