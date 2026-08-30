from __future__ import annotations

import base64
from io import BytesIO
import unittest
import zlib

from openpyxl import Workbook

from pipeline.internal.capital_weekly.context.metal_inventories import (
    comex_schema_signature,
    parse_comex_stocks,
    parse_usgs_mcs_pdf,
    parse_usgs_mcs_text,
)


COPPER_SPEC = {
    "commodity_code": "COPPER_COMEX",
    "commodity_family": "copper",
    "commodity_title": "COPPER - HIGH GRADE",
    "expected_sheet": "Daily Metal Stocks Report",
    "expected_unit": "Short Tons",
    "location_header": "DELIVERY POINT",
    "registered_total_label": "Total Registered (warranted)",
    "eligible_total_label": "Total Eligible (non-warranted)",
    "combined_total_label": "TOTAL COPPER",
}
GOLD_SPEC = {
    "commodity_code": "GOLD_COMEX",
    "commodity_family": "gold",
    "commodity_title": "GOLD",
    "expected_sheet": "Daily Metal Stocks Report",
    "expected_unit": "Troy Ounce",
    "location_header": "DEPOSITORY",
    "registered_total_label": "TOTAL REGISTERED",
    "eligible_total_label": "TOTAL ELIGIBLE",
    "combined_total_label": "COMBINED TOTAL",
}

# A compact deterministic BIFF8 workbook with the same schema as comex_xlsx().
# It was generated once from that fixture and compressed so tests need no writer.
COPPER_BIFF8 = zlib.decompress(base64.b64decode(
    "eNrtWG1sU1UYfm/bdbej68e+GTCvKLCPbsBQGChspb2u1W1d2ssGBiNl3I1KaZeuKCMxFtF/"
    "QDTE6A8SQ8TExPgVE/2hxu0fIRr9ww+NJoNfGvxRjcaQyOp73nt61867ZR1oou60zz33fc9"
    "5zud7Pt779VfumUsfNF6HeeEhMMNszgbWAp2AWJUXXAAmrpvN5XJ5dW4l/KvCbR6zObTg/J"
    "Uh2JyXI0SEDVHB59jO45Xw3wlhSOIvDRLIkMA4BZNQSqhDiyksbymc2SXmW2pYqX/59bP9m"
    "+3jZsRi65/t+2z9VyIcCKd2BIAbUYWoRtQgaskmAOoRDYjViEbEGsRaxDpEE+IehIS4F7Ee"
    "cR/ifsQGxEbEJkQzogXRimhDeBDtiA7EZsQW3pdOHjM8iO/bETu4bifGu1b2rQWDTcRZtJb"
    "BJ5VfsCmnub+OFvG+ZZrs4gbiIIwz2/BF4yPSPxX2UhuiAmvDFBrpbnwT4CJqHfAeaT+l52"
    "6yPgw9IHVxi95v6qG2n6Pneno6gPE/Js63pNkKnfAdWflLfBG4hT6I4S54Ak6ibbL3w6Dij"
    "pgmrYq6XmQ1whW25p4jFlsxghfzxCAK8TtMbSpqjwDlgo9SWDtYriaLHS6zldmrJtRUND5D"
    "K+4y/JaTClb3lMT0rAjS/7o0valEPfzv9Jdx7IzGuU3Ln52vb1lA71lA37qA3maoL8u3R5z"
    "TXzDhKs1AjsWujJlid8ZKcVXGQnF1ppzimkxZ7hCthhdxD36CrvmrwR+NxSelfjUdjUuRdHL"
    "k2IQUVseTqfRZ3ACG8DctiAgA5ZwAN+Fxdgxkw1S9Ncu2e7ZP5N2CQmyDVVm2UD/H9opwWs"
    "BjRZwSevDZA00/u7SDRcQDRsSDQ/wdz5Y/4BWhgR8oGPtC/f0hf1A5IMn7fQHvQK/skYIDv"
    "g48P/plxdsnDXvDciC0LyJLESXkeyyCkVcJRpSgL1LN6IODclhqlwLB3oDUG/b6ZTyltL5J/"
    "mha3SV1be7E/5bO7XjcRY4yvZJMTGDV3pF07OlYenIu3w7Kh8egX+4LDsnhA9JgKDigIHEwL"
    "A9JSggbhPMSln0yJvuxs8NBJeAPe4cHMM+ArEhaF1Dw+h/dF1H65QEFT1giIt3vPYDnsLdvM"
    "OClVo7FJtJqSj0iNT8TTaWiibR6pAUbJsdjY7HDcVVqTiQT7XNJyFWSOIU4F3txbNZyUTIu"
    "qCmfvFBx9nzDtEHM4X2gFt60aPeBEbEWMbcqanlsYpcG7XBxFR0udrSPI2g3bDt2k6240CL"
    "afWtv3tje2P0k6TNzVg0bmFGDRziNKdOWVmK00fN5ypu/dJwh3Qv03Kjzf+jeVPDerL9nu1"
    "sK3i/hIWHDVjFbFbAEj+ARGuzWM9OHZrrzsQD7MEWEZ6nbIKRsTt5TB12cVpPAyhRNLtLgUt"
    "e2ecFJOquBrpyuWsU60UBXQWUX6+wG5VUa5HMY6JwGOrdBeVUG+aoNdDUGujoDXb2BrmGe7j"
    "b21qKbEpOsJAlcKifJpEtWlMxcEinNokssrYxLFZRm1SUTSuW6ZEZJ1CULSjZdKkOpQpdYma"
    "t0qRwlO5fsVEMllypJYqP/GW7ulVifk66yTmq70bOVZp5xHcR1cq6DuOFFuWGd6ySui3Odx"
    "D21KPeUznUT1817UFXUgyoqqW7Rkur0kqqLelBN3D2Lcvfo3JqiHtSU1Ps64lZxbl1J9dYTt"
    "5pz64l7flHueZ2rnVU1nNtA3KuLcq8S9ycYZJpsQNsJsyI/MZkRO8Gc3UJ5Ldlm0tuyLgOHz"
    "gQVxGMcZu17TFXwkVW7FQM8TDsbc72ctE7yFWAYw+N2TCwsSMg7kRVL8z0F9KzMFZoP6hJN"
    "2k1ES3qE+Xz9sZFUciI5mpbkkyNqXNq5o12JHlbjcZVW4t7Y6GjXnfi+gnbfWHaYzQFt+Qb"
    "lwsyLr/9yK3TU9fbLIrRt+vAbNhWvcd9Y4IPLtpoA95MP8mE7yv3lNPeZM9xvPst97R9vaz"
    "4w4yTH1cT45Mk4cL/YxP3anoK2/OW9NvvOG13fC4Xvd7v/TGO69uW1ix1rXBdexf57br3rZ"
    "x7LPN1B7ufnL3qugj728e8KwH19lt7F03sKvj2wsfKOjw+pqYlYMlE8Rmx8t3VshZXw94S7"
    "+f2P5n/eHmDEYTYx5dDeh7H2FBxDD5e141jJ7WeHkonbSa6E72rr7Pl17sN6j8M4hLANTy"
    "2rfnPBN7ilcJg1B/V9JoQevrrs+WPXmFK//23j3+roCwRE4AT2/zhEae6DaAWjNCdMw758J"
    "FGzcGjm9ZeVMP4PIK7o9fuxhhFqg0oWWFp7upbRf7YHvbXy/ZfCnyYRW70="
))


def simple_pdf(lines: list[str]) -> bytes:
    def escape(value: str) -> str:
        return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    commands = ["BT", "/F1 10 Tf", "72 720 Td", "12 TL"]
    for index, line in enumerate(lines):
        commands.append(f"({escape(line)}) Tj")
        if index + 1 < len(lines):
            commands.append("T*")
    commands.append("ET")
    stream = "\n".join(commands).encode("latin-1")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        (
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            b"/Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>"
        ),
        b"<< /Length "
        + str(len(stream)).encode("ascii")
        + b" >>\nstream\n"
        + stream
        + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    output = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = []
    for number, item in enumerate(objects, 1):
        offsets.append(len(output))
        output.extend(f"{number} 0 obj\n".encode("ascii"))
        output.extend(item)
        output.extend(b"\nendobj\n")
    xref = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref}\n%%EOF\n"
        ).encode("ascii")
    )
    return bytes(output)


USGS_GOLD_PDF = simple_pdf(
    [
        "GOLD",
        "(Data in metric tons,1 gold content, unless otherwise specified)",
        "World Mine Production and Reserves:",
        "Mine production Reserves10",
        "2024 2025e",
        "World total (rounded) 3,280 3,300 66,000",
        "U.S. Geological Survey, Mineral Commodity Summaries, February 2026",
    ]
)


def comex_xlsx(
    *,
    unit: str = "Short Tons",
    location_header: str = "DELIVERY POINT",
    alpha_total: float = 30,
    eligible_total: float = 35,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Metal Stocks Report"
    sheet["A5"] = "COMMODITY EXCHANGE, INC."
    sheet["A7"] = "METAL WAREHOUSE STOCKS STATISTICS"
    sheet["A8"] = "COPPER - HIGH GRADE"
    sheet["G8"] = "Report Date: 8/28/2026"
    sheet["A9"] = unit
    sheet["G9"] = "Activity Date: 8/27/2026"
    headers = (
        location_header,
        "",
        "PREV TOTAL",
        "RECEIVED",
        "WITHDRAWN",
        "NET CHANGE",
        "ADJUSTMENT",
        "TOTAL TODAY",
    )
    for column, value in enumerate(headers, 1):
        sheet.cell(11, column, value)
    rows = {
        13: ("ALPHA",),
        14: ("Registered (warranted)", "", 10, 0, 0, 0, 0, 10),
        15: ("Eligible (non-warranted)", "", 20, 0, 0, 0, 0, 20),
        16: ("Total", "", 30, 0, 0, 0, 0, alpha_total),
        18: ("BETA",),
        19: ("Registered (warranted)", "", 5, 0, 0, 0, 0, 5),
        20: ("Eligible (non-warranted)", "", 15, 0, 0, 0, 0, 15),
        21: ("Total", "", 20, 0, 0, 0, 0, 20),
        23: ("Total Registered (warranted)", "", 15, 0, 0, 0, 0, 15),
        24: (
            "Total Eligible (non-warranted)",
            "",
            35,
            0,
            0,
            0,
            0,
            eligible_total,
        ),
        25: ("TOTAL COPPER", "", 50, 0, 0, 0, 0, 50),
    }
    for row_number, values in rows.items():
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def gold_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Metal Stocks Report"
    sheet["A8"] = "GOLD"
    sheet["G8"] = "Report Date: 8/28/2026"
    sheet["A9"] = "Troy Ounce"
    for column, value in enumerate(
        (
            "DEPOSITORY",
            "",
            "PREV TOTAL",
            "RECEIVED",
            "WITHDRAWN",
            "NET CHANGE",
            "ADJUSTMENT",
            "TOTAL TODAY",
        ),
        1,
    ):
        sheet.cell(11, column, value)
    rows = {
        13: ("ALPHA",),
        14: ("Registered", "", 100, 0, 0, 0, 0, 100),
        15: ("Pledged", "", 25, "", "", "", "", 25),
        16: ("Eligible", "", 50, 0, 0, 0, 0, 50),
        17: ("Total", "", 150, 0, 0, 0, 0, 150),
        19: ("TOTAL REGISTERED", "", 100, 0, 0, 0, 0, 100),
        20: ("TOTAL PLEDGED", "", 25, "", "", "", "", 25),
        21: ("TOTAL ELIGIBLE", "", 50, 0, 0, 0, 0, 50),
        22: ("COMBINED TOTAL", "", 150, 0, 0, 0, 0, 150),
    }
    for row_number, values in rows.items():
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class ComexStockParserTests(unittest.TestCase):
    def test_preserves_location_rows_and_reconciled_exchange_totals(self):
        rows = parse_comex_stocks(comex_xlsx(), COPPER_SPEC)

        self.assertEqual(len(rows), 9)
        exchange = [row for row in rows if row["scope"] == "exchange"]
        self.assertEqual(
            [(row["inventory_type"], row["value"]) for row in exchange],
            [("registered", 15.0), ("eligible", 35.0), ("total", 50.0)],
        )
        self.assertTrue(all(row["report_date"].isoformat() == "2026-08-28" for row in rows))
        self.assertTrue(all(row["unit"] == "Short Tons" for row in rows))
        self.assertEqual(
            {(row["location"], row["inventory_type"]) for row in rows[:6]},
            {
                ("ALPHA", "registered"),
                ("ALPHA", "eligible"),
                ("ALPHA", "total"),
                ("BETA", "registered"),
                ("BETA", "eligible"),
                ("BETA", "total"),
            },
        )

    def test_rejects_changed_header_or_source_unit(self):
        with self.assertRaisesRegex(ValueError, "COMEX header mismatch"):
            parse_comex_stocks(
                comex_xlsx(location_header="WAREHOUSE"),
                COPPER_SPEC,
            )
        with self.assertRaisesRegex(ValueError, "COMEX unit mismatch"):
            parse_comex_stocks(comex_xlsx(unit="Metric Tons"), COPPER_SPEC)

    def test_rejects_non_finite_values_and_any_total_mismatch(self):
        with self.assertRaisesRegex(ValueError, "finite"):
            parse_comex_stocks(comex_xlsx(eligible_total=float("nan")), COPPER_SPEC)
        with self.assertRaisesRegex(ValueError, "ALPHA.*does not reconcile"):
            parse_comex_stocks(comex_xlsx(alpha_total=31), COPPER_SPEC)

    def test_reports_ooxml_schema_signature(self):
        signature = comex_schema_signature(comex_xlsx(), COPPER_SPEC)
        self.assertRegex(signature, r"^ooxml-xlsx:sha256:[0-9a-f]{64}$")

    def test_gold_pledged_subset_is_not_double_counted_in_total(self):
        rows = parse_comex_stocks(gold_xlsx(), GOLD_SPEC)

        self.assertFalse(any(row["inventory_type"] == "pledged" for row in rows))
        self.assertEqual(
            [(row["inventory_type"], row["value"]) for row in rows[-3:]],
            [("registered", 100.0), ("eligible", 50.0), ("total", 150.0)],
        )

    def test_parses_the_verified_current_ole2_biff8_container(self):
        rows = parse_comex_stocks(COPPER_BIFF8, COPPER_SPEC)

        self.assertEqual(
            [(row["inventory_type"], row["value"]) for row in rows[-3:]],
            [("registered", 15.0), ("eligible", 35.0), ("total", 50.0)],
        )
        self.assertRegex(
            comex_schema_signature(COPPER_BIFF8, COPPER_SPEC),
            r"^ole2-biff8:sha256:[0-9a-f]{64}$",
        )


class UsgsStructuralParserTests(unittest.TestCase):
    def test_parses_current_world_copper_production_and_reserves(self):
        text = """
COPPER
(Data in thousand metric tons, copper content, unless otherwise specified)
World Mine and Refinery Production and Reserves:
Mine production Refinery production Reserves6
2024 2025e 2024 2025e
World total (rounded) 23,000 23,000 27,600 29,000 980,000
U.S. Geological Survey, Mineral Commodity Summaries, February 2026
"""
        rows = parse_usgs_mcs_text(
            text,
            {
                "commodity_code": "COPPER_COMEX",
                "commodity_family": "copper",
                "commodity_title": "COPPER",
                "expected_unit": "thousand metric tons, copper content",
                "table_kind": "mine_refinery_reserves",
                "reference_year": "2025",
                "publication_month": "February 2026",
            },
        )

        self.assertEqual(
            [(row["measurement"], row["value"], row["unit"]) for row in rows],
            [
                ("mine_production", 23_000.0, "thousand metric tons, copper content"),
                ("reserves", 980_000.0, "thousand metric tons, copper content"),
            ],
        )
        self.assertTrue(all(row["reference_period"] == "2025" for row in rows))

    def test_ignores_the_gold_unit_footnote_marker_but_not_unit_text(self):
        text = """
GOLD
(Data in metric tons,1 gold content, unless otherwise specified)
World Mine Production and Reserves:
Mine production Reserves10
2024 2025e
World total (rounded) 3,280 3,300 66,000
U.S. Geological Survey, Mineral Commodity Summaries, February 2026
"""
        rows = parse_usgs_mcs_text(
            text,
            {
                "commodity_code": "GOLD_COMEX",
                "commodity_family": "gold",
                "commodity_title": "GOLD",
                "expected_unit": "metric tons, gold content",
                "table_kind": "mine_reserves",
                "reference_year": "2025",
                "publication_month": "February 2026",
            },
        )

        self.assertEqual(
            [(row["measurement"], row["value"]) for row in rows],
            [("mine_production", 3300.0), ("reserves", 66000.0)],
        )

    def test_extracts_the_structural_table_from_pdf_bytes(self):
        rows = parse_usgs_mcs_pdf(
            USGS_GOLD_PDF,
            {
                "commodity_code": "GOLD_COMEX",
                "commodity_family": "gold",
                "commodity_title": "GOLD",
                "expected_unit": "metric tons, gold content",
                "table_kind": "mine_reserves",
                "reference_year": "2025",
                "publication_month": "February 2026",
            },
        )

        self.assertEqual([row["value"] for row in rows], [3300.0, 66000.0])

    def test_rejects_changed_usgs_unit_or_incomplete_world_total(self):
        base = """
GOLD
(Data in metric tons, gold content, unless otherwise specified)
World Mine Production and Reserves:
Mine production Reserves10
2024 2025e
World total (rounded) 3,280 3,300 66,000
U.S. Geological Survey, Mineral Commodity Summaries, February 2026
"""
        spec = {
            "commodity_code": "GOLD_COMEX",
            "commodity_family": "gold",
            "commodity_title": "GOLD",
            "expected_unit": "metric tons, gold content",
            "table_kind": "mine_reserves",
            "reference_year": "2025",
            "publication_month": "February 2026",
        }
        with self.assertRaisesRegex(ValueError, "USGS unit mismatch"):
            parse_usgs_mcs_text(
                base.replace("metric tons, gold content", "troy ounces"),
                spec,
            )
        with self.assertRaisesRegex(ValueError, "world total"):
            parse_usgs_mcs_text(base.replace(" 66,000", ""), spec)


if __name__ == "__main__":
    unittest.main()
