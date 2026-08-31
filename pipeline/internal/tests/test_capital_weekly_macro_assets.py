from datetime import date
from dataclasses import replace
from io import StringIO
import json
import io
import os
from pathlib import Path
import re
import tempfile
import unittest
from unittest.mock import patch

import pandas as pd
import requests
from openpyxl import Workbook

from pipeline.internal.capital_weekly import macro_assets
from pipeline.internal.capital_weekly import macro_assets as macro_assets_module
from pipeline.internal.capital_weekly.context.eia_commodities import EiaBatchSpec
from pipeline.internal.capital_weekly.official_http import (
    OfficialHttpResponse,
    OfficialHttpTrace,
)
from pipeline.internal.capital_weekly.macro_assets import (
    _atomic_write_bytes,
    _session,
    _parse_fred_csv,
    _parse_chinabond_json,
    _parse_sina_fx_day_kline,
    _parse_treasury_csv,
    _parse_yahoo_chart,
    _parse_boc_json, _parse_boe_csv, _parse_ecb_csv, _parse_hkma_json,
    _parse_nyfed_json, _parse_snb_csv, _parse_boj_json, _carry_forward_business_daily,
    _parse_boj_policy_text, _parse_boj_policy_summary_text,
    _parse_boj_policy_candidates,
    _parse_boj_policy_announcement, _parse_pboc_omo_announcement,
    _parse_rbi_current_rate, _parse_rbi_history,
    _parse_rate_xlsx,
    _discover_world_bank_monthly_url,
    _fetch_config_history,
    MacroAssetBundle,
    MacroAssetConfig,
    align_series_histories,
    align_curve_spread,
    calculate_five_year_five_year,
    fetch_macro_assets,
    fetch_macro_asset_bundle,
    load_macro_asset_universe,
)


class MacroAssetUniverseTests(unittest.TestCase):
    def test_world_bank_transport_uses_exact_policy_and_does_not_retry_parser(self):
        page = (
            '<a href="https://thedocs.worldbank.org/official/'
            'CMO-Historical-Data-Monthly.xlsx">Monthly prices</a>'
        ).encode()
        workbook = b"invalid-world-bank-workbook\x00\xff"
        calls = []

        def fake_get(_session, url, **kwargs):
            calls.append((url, kwargs))
            body = page if len(calls) == 1 else workbook
            return OfficialHttpResponse(
                body=body,
                url=url,
                headers={},
                trace=OfficialHttpTrace(2, 1, [503, 200], url),
            )

        config = replace(
            self._config("world_bank_pink_sheet", "Gold"),
            source_url="https://www.worldbank.org/en/research/commodity-markets",
            level_unit="$/toz",
        )
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        policy = macro_assets_module.load_commodity_http_policies()[
            "world_bank_pink_sheet"
        ].policy

        with patch.object(macro_assets_module, "official_get", side_effect=fake_get), patch.object(
            macro_assets_module,
            "parse_world_bank_monthly_prices",
            side_effect=ValueError("World Bank parser rejected fixture"),
        ) as parser:
            with self.assertRaisesRegex(ValueError, "parser rejected"):
                _fetch_config_history(config, session)

        self.assertEqual(len(calls), 2)
        self.assertEqual(parser.call_count, 1)
        self.assertTrue(all(kwargs["policy"] == policy for _, kwargs in calls))
        self.assertTrue(all(kwargs["audit_secrets"] == () for _, kwargs in calls))

    def test_macro_eia_metadata_requires_explicit_total(self):
        spec = EiaBatchSpec(
            route="petroleum/pri/spt",
            facets={"series": ("RWTC",)},
            frequency="daily",
            start="2026-08-01",
            end="2026-08-23",
            page_length=1,
        )
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        http = macro_assets_module.load_commodity_http_policies()["eia"]
        client = macro_assets_module._MacroEiaClient(
            session, "fixture-key", http
        )
        body = json.dumps(
            {"response": {"facets": [{"id": "RWTC"}]}}
        ).encode()

        with patch.object(
            macro_assets_module, "_official_macro_get", return_value=body
        ):
            with self.assertRaisesRegex(ValueError, "total"):
                client.fetch_metadata(
                    spec,
                    {"RWTC": {"facets": {"series": "RWTC"}}},
                )

    def test_fred_overrides_the_browser_user_agent_with_requests_default(self):
        self.assertEqual(
            _session().headers["User-Agent"],
            "Mozilla/5.0 (capital-weekly research)",
        )
        config = self._config("fred", "DFEDTARL")
        session = unittest.mock.Mock(
            _macro_attempt_trace=[],
            _macro_raw_parts=[],
        )
        response = unittest.mock.Mock(
            content=(
                b"observation_date,DFEDTARL\n"
                b"2026-07-10,3.50\n"
            ),
            text="observation_date,DFEDTARL\n2026-07-10,3.50\n",
        )
        response.raise_for_status.return_value = None
        session.get.return_value = response

        _fetch_config_history(
            config,
            session,
            as_of_date=date(2026, 8, 23),
        )

        self.assertEqual(
            session.get.call_args.kwargs["headers"],
            {"User-Agent": requests.utils.default_user_agent()},
        )

    def _assert_policy_universe_semantics(self, universe):
        for item in universe:
            if item.series_code != "ECB_MLF":
                semantic_fields = (
                    item.series_code, item.name_cn, item.name_en,
                    item.provider, item.provider_symbol, item.source,
                )
                self.assertFalse(any(
                    re.search(r"(?:^|[^A-Z])MLF(?:$|[^A-Z])", field.upper())
                    or "MEDIUM-TERM LENDING FACILITY" in field.upper()
                    for field in semantic_fields
                ))
            self.assertNotEqual(item.series_code.strip().upper(), "DR007")
            self.assertNotEqual(item.provider_symbol.strip().upper(), "DR007")

        approved_fdr007 = [
            item for item in universe
            if item.series_code == "CNY_FDR007" and item.provider_symbol == "FDR007"
        ]
        self.assertEqual(len(approved_fdr007), 1)
        self.assertEqual(
            sum(item.provider_symbol.strip().upper() == "FDR007" for item in universe),
            1,
        )

    def _config(self, provider="fred", symbol="TEST"):
        return MacroAssetConfig(
            "fixed_income", "x", "TEST", "测", "Test", provider, symbol,
            "Source", "https://generic.test", "daily", "percent", "bp", 1, "",
        )

    def test_strict_records_preserves_change_precision_below_1e_12(self):
        from pipeline.internal.scripts.fetch_macro_assets import strict_records

        value = 0.0123456789012345
        records = strict_records(pd.DataFrame({"daily_change": [value]}))

        self.assertLess(abs(records[0]["daily_change"] - value), 1e-12)

    def test_yahoo_chart_rejects_mismatched_timestamp_and_close_lengths(self):
        fixture = json.dumps({"chart": {"result": [{
            "timestamp": [1783641600, 1783728000],
            "indicators": {"quote": [{"close": [100.5]}]},
        }]}})

        with self.assertRaisesRegex(ValueError, "length"):
            _parse_yahoo_chart(fixture)

    def test_sina_fx_day_kline_parses_daily_usd_cnh_closes(self):
        fixture = (
            "/*<script>location.href='//sina.com';</script>*/\n"
            'var_fx_susdcnh=("2026-07-23,6.7800,6.7700,6.7900,6.7750,|'
            '2026-07-24,6.7760,6.7650,6.7820,6.7716,");'
        )

        self.assertEqual(
            _parse_sina_fx_day_kline(fixture, "fx_susdcnh"),
            [
                {"date": date(2026, 7, 23), "value": 6.775},
                {"date": date(2026, 7, 24), "value": 6.7716},
            ],
        )

    def test_sina_fx_day_kline_rejects_wrong_variable_and_duplicate_dates(self):
        with self.assertRaisesRegex(ValueError, "variable"):
            _parse_sina_fx_day_kline(
                'var_fx_susdcny=("2026-07-24,6.7,6.6,6.8,6.7,");',
                "fx_susdcnh",
            )
        with self.assertRaisesRegex(ValueError, "duplicate"):
            _parse_sina_fx_day_kline(
                'var_fx_susdcnh=("2026-07-24,6.7,6.6,6.8,6.7,|'
                '2026-07-24,6.7,6.6,6.8,6.7,");',
                "fx_susdcnh",
            )

    def test_sina_fx_day_kline_rejects_non_positive_close(self):
        with self.assertRaisesRegex(ValueError, "positive"):
            _parse_sina_fx_day_kline(
                'var_fx_susdcnh=("2026-07-24,6.7,6.6,6.8,0,");',
                "fx_susdcnh",
            )

    def test_official_policy_json_and_csv_fixtures(self):
        self.assertEqual(_parse_nyfed_json(json.dumps({"refRates": [{"type": "EFFR", "effectiveDate": "2026-07-10", "percentRate": 3.64}, {"type": "SOFRAI", "effectiveDate": "2026-07-10"}]}), "EFFR"), [{"date": date(2026, 7, 10), "value": 3.64}])
        self.assertEqual(_parse_ecb_csv("TIME_PERIOD,OBS_VALUE\n2026-07-10,2.00\n"), [{"date": date(2026, 7, 10), "value": 2.0}])
        self.assertEqual(_parse_boe_csv("DATE,IUDBEDR,IUDSOIA\n10 Jul 2026,4.25,3.97\n", "IUDBEDR"), [{"date": date(2026, 7, 10), "value": 4.25}])
        self.assertEqual(_parse_hkma_json(json.dumps({"header": {"success": True}, "result": {"records": [{"end_of_date": "2026-07-10", "disc_win_base_rate": "4.75"}]}}), "disc_win_base_rate"), [{"date": date(2026, 7, 10), "value": 4.75}])
        self.assertEqual(_parse_boc_json(json.dumps({"observations": [{"d": "2026-07-10", "V39079": {"v": "2.75"}}]}), "V39079"), [{"date": date(2026, 7, 10), "value": 2.75}])
        self.assertEqual(_parse_snb_csv("CubeId;snboffzisa\nDate;D0;Value\n2026-06-30;LZ;0.25\n", "LZ"), [{"date": date(2026, 6, 30), "value": 0.25}])

    def test_boj_api_accepts_real_nested_values_shape(self):
        fixture = json.dumps({"RESULTSET": [{
            "SERIES_CODE": "STRDCLUCON",
            "VALUES": {"SURVEY_DATES": [20260709, 20260710], "VALUES": ["0.476", "0.477"]},
        }]})
        self.assertEqual(_parse_boj_json(fixture), [
            {"date": date(2026, 7, 9), "value": 0.476},
            {"date": date(2026, 7, 10), "value": 0.477},
        ])

    def test_snb_parser_accepts_real_quoted_monthly_shape(self):
        fixture = '\ufeff"CubeId";"snboffzisa"\n"PublishingDate";"2026-06-22 09:00"\n\n"Date";"D0";"Value"\n"2026-05";"LZ";"0"\n'
        self.assertEqual(_parse_snb_csv(fixture, "LZ"), [
            {"date": date(2026, 5, 1), "value": 0.0},
        ])

    def test_snb_parser_requires_selector_column_and_exact_match(self):
        with self.assertRaisesRegex(ValueError, "D0"):
            _parse_snb_csv('"Date";"Value"\n"2026-05";"0"\n', "LZ")
        with self.assertRaisesRegex(ValueError, "LZ"):
            _parse_snb_csv('"Date";"D0";"Value"\n"2026-05";"OTHER";"0"\n', "LZ")

    def test_boj_policy_text_uses_effective_date_from_real_pdf_shape(self):
        fixture = """June 16, 2026 Bank of Japan
        The Bank will encourage the uncollateralized overnight call rate to remain at around 1.0 percent.
        The new guideline for money market operations will be effective from June 17, 2026.
        """
        self.assertEqual(_parse_boj_policy_text(fixture), {
            "date": date(2026, 6, 17), "value": 1.0,
        })

    def test_boj_policy_summary_accepts_real_2025_shape(self):
        fixture = 'Decision at the December 2025 MPM Short-term interest rate : raised to "around 0.75%" (uncollateralized overnight call rate)'
        self.assertEqual(_parse_boj_policy_summary_text(fixture, date(2025, 12, 19)), {
            "date": date(2025, 12, 19), "value": 0.75,
        })

    def test_boj_newest_malformed_candidate_fails_instead_of_carrying_old_rate(self):
        candidates = [
            ("https://www.boj.or.jp/en/mopo/mpmdeci/mpr_2025/k251219b.pdf", b"old"),
            ("https://www.boj.or.jp/en/mopo/mpmdeci/mpr_2026/k260616b.pdf", b"new-malformed"),
        ]
        with patch("pipeline.internal.capital_weekly.macro_assets._parse_boj_policy_summary_pdf") as parser:
            parser.side_effect = [
                {"date": date(2025, 12, 19), "value": 0.75},
                ValueError("malformed"),
            ]
            with self.assertRaisesRegex(ValueError, "k260616b"):
                _parse_boj_policy_candidates(candidates)

    def test_event_parsers_bind_date_and_value_within_one_official_record(self):
        boj = '<time>2026-06-17</time><p>The Bank will encourage the uncollateralized overnight call rate to remain at around 0.5 percent.</p>'
        pboc = '<div>2026年7月10日</div><table><tr><td>期限</td><td>7天</td></tr><tr><td>操作利率</td><td>1.40%</td></tr></table>'
        rbi_current = '<div>As on 10 July 2026</div><tr><td>Policy Repo Rate</td><td>5.50%</td></tr>'
        rbi_history = '<table><tr><th>Effective Date</th><th>Repo</th></tr><tr><td>09/07/2026</td><td>5.50</td></tr></table>'
        self.assertEqual(_parse_boj_policy_announcement(boj), {"date": date(2026, 6, 17), "value": 0.5})
        self.assertEqual(_parse_pboc_omo_announcement(pboc), {"date": date(2026, 7, 10), "value": 1.4})
        self.assertEqual(_parse_rbi_current_rate(rbi_current), {"date": date(2026, 7, 10), "value": 5.5})
        self.assertEqual(_parse_rbi_history(rbi_history), [{"date": date(2026, 7, 9), "value": 5.5}])

    def test_chinabond_accepts_only_successful_official_envelope(self):
        payload = json.dumps({
            "flag": "0",
            "heList": [{
                "workTime": "2026-08-28",
                "twoYear": "1.24",
                "fiveYear": "1.39",
                "tenYear": "1.68",
                "thirtyYear": "2.13",
            }],
        })
        self.assertEqual(
            _parse_chinabond_json(payload, "tenYear"),
            [{"date": date(2026, 8, 28), "value": 1.68}],
        )
        for invalid in (
            {"flag": "1", "heList": []},
            {"flag": "0"},
            {"flag": "0", "heList": {}},
        ):
            with self.subTest(invalid=invalid):
                with self.assertRaises(ValueError):
                    _parse_chinabond_json(json.dumps(invalid), "tenYear")

    def test_chinabond_posts_current_history_endpoint_and_parameters(self):
        response = unittest.mock.Mock(
            content=b'{"flag":"0","heList":[]}',
            text='{"flag":"0","heList":[]}',
        )
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(
            _macro_attempt_trace=[],
            _macro_raw_parts=[],
            post=unittest.mock.Mock(return_value=response),
        )
        with patch(
            "pipeline.internal.capital_weekly.macro_assets._parse_chinabond_json",
            return_value=[],
        ):
            _fetch_config_history(
                self._config("china_bond", "10Y"),
                session,
                as_of_date=date(2026, 8, 30),
            )
        url = session.post.call_args.args[0]
        self.assertIn("/cbweb-czb-web/czb/historyQuery?", url)
        self.assertIn("gjqx=2,5,10,30", url)
        self.assertIn("locale=en_US", url)
        self.assertIn("qxmc=1", url)

    def test_pboc_parser_rejects_non_seven_day_operation(self):
        html = '<div>2026年7月10日</div><table><tr><td>期限</td><td>14天</td></tr><tr><td>操作利率</td><td>1.55%</td></tr></table>'
        with self.assertRaisesRegex(ValueError, "7-day"):
            _parse_pboc_omo_announcement(html)

    def test_rba_workbook_fixture_selects_exact_series_id(self):
        frame = pd.DataFrame([[None, "FIRMMCRT", "FIRMMCRI"], [date(2026, 7, 9), 3.6, 3.58]])
        stream = io.BytesIO()
        with pd.ExcelWriter(stream, engine="openpyxl") as writer:
            frame.to_excel(writer, index=False, header=False)
        self.assertEqual(_parse_rate_xlsx(stream.getvalue(), "FIRMMCRT"), [{"date": date(2026, 7, 9), "value": 3.6}])

    def test_fred_dispatch_bounds_request_dates(self):
        config = self._config("fred", "DFEDTARL")
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        response = unittest.mock.Mock(content=b"observation_date,DFEDTARL\n2026-07-10,3.50\n", text="observation_date,DFEDTARL\n2026-07-10,3.50\n")
        response.raise_for_status.return_value = None
        session.get.return_value = response
        with patch("pipeline.internal.capital_weekly.macro_assets.date") as mocked_date:
            mocked_date.today.return_value = date(2026, 7, 13)
            _fetch_config_history(config, session)
        url = session.get.call_args.args[0]
        self.assertIn("cosd=2025-01-09", url)
        self.assertIn("coed=2026-07-13", url)

    def test_eia_v2_dispatch_uses_route_symbol_and_keeps_key_out_of_provenance(self):
        config = replace(
            self._config("eia_v2", "RWTC"),
            provider_route="petroleum/pri/spt",
            level_unit="Dollars per Barrel",
            source_description="Cushing WTI Spot Price FOB",
        )
        text = json.dumps({"response": {"total": 1, "data": [
            {
                "period": "2026-07-10",
                "series": "RWTC",
                "series-description": "Cushing WTI Spot Price FOB",
                "unit": "Dollars per Barrel",
                "value": "68.25",
            }
        ]}})
        response = unittest.mock.Mock(
            content=text.encode(), text=text, status_code=200, headers={}
        )
        metadata_text = json.dumps(
            {"response": {
                "totalFacets": 1,
                "facets": [{"id": "RWTC", "name": "WTI"}],
            }}
        )
        metadata_response = unittest.mock.Mock(
            content=metadata_text.encode(), text=metadata_text,
            status_code=200, headers={},
        )
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.side_effect = [metadata_response, response]

        with patch.dict(os.environ, {"EIA_API_KEY": "test-key"}):
            history, raw, provenance = _fetch_config_history(
                config,
                session,
                as_of_date=date(2026, 7, 11),
            )

        self.assertEqual(history[0]["value"], 68.25)
        self.assertEqual(raw, metadata_text.encode() + b"\n" + text.encode())
        self.assertEqual(
            provenance,
            "https://api.eia.gov/v2/petroleum/pri/spt/data/",
        )
        self.assertNotIn("test-key", provenance)
        self.assertEqual(
            session.get.call_args.kwargs["params"]["facets[series][]"],
            ["RWTC"],
        )
        self.assertEqual(session.get.call_args.kwargs["params"]["api_key"], "test-key")

    def test_world_bank_workbook_is_discovered_downloaded_once_and_parsed_once(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Monthly Prices"
        sheet.append(["Monthly Prices"])
        sheet.append(["Date", "Gold", "Copper"])
        sheet.append(["", "$/toz", "$/mt"])
        sheet.append(["2026M06", 2340.5, 9510.0])
        stream = io.BytesIO()
        workbook.save(stream)
        workbook_bytes = stream.getvalue()
        page = (
            '<a href="https://thedocs.worldbank.org/official/'
            'CMO-Historical-Data-Monthly.xlsx">Monthly prices</a>'
        )
        page_response = unittest.mock.Mock(
            content=page.encode(), text=page, status_code=200, headers={}
        )
        page_response.raise_for_status.return_value = None
        workbook_response = unittest.mock.Mock(
            content=workbook_bytes, text="", status_code=200, headers={}
        )
        workbook_response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.side_effect = [page_response, workbook_response]
        session._macro_world_bank_columns = {"Gold": "$/toz", "Copper": "$/mt"}
        gold = replace(
            self._config("world_bank_pink_sheet", "Gold"),
            source_url="https://www.worldbank.org/en/research/commodity-markets",
            level_unit="$/toz",
        )
        copper = replace(gold, provider_symbol="Copper", level_unit="$/mt")

        gold_history, gold_raw, gold_url = _fetch_config_history(gold, session)
        session._macro_attempt_trace = []
        session._macro_raw_parts = []
        copper_history, copper_raw, copper_url = _fetch_config_history(copper, session)

        self.assertEqual(gold_history[0]["value"], 2340.5)
        self.assertEqual(copper_history[0]["value"], 9510.0)
        self.assertEqual(gold_raw, workbook_bytes)
        self.assertEqual(copper_raw, workbook_bytes)
        self.assertEqual(gold_url, copper_url)
        self.assertEqual(session.get.call_count, 2)

    def test_historical_treasury_bundle_filters_source_rows_after_as_of_date(self):
        config = replace(
            self._config("us_treasury", "10-year"),
            series_code="UST10Y",
        )
        session = unittest.mock.Mock(headers={})

        def response_for(url, **_kwargs):
            text = (
                "Date,10 Yr\n"
                if "/2025/" in url
                else (
                    "Date,10 Yr\n"
                    "07/24/2026,4.10\n"
                    "07/31/2026,4.20\n"
                    "08/03/2026,9.99\n"
                )
            )
            response = unittest.mock.Mock(content=text.encode(), text=text)
            response.raise_for_status.return_value = None
            return response

        session.get.side_effect = response_for
        with patch.object(
            macro_assets_module,
            "load_macro_asset_universe",
            return_value=[config],
        ), patch.object(
            macro_assets_module,
            "_session",
            return_value=session,
        ):
            try:
                bundle = fetch_macro_asset_bundle(as_of_date=date(2026, 8, 2))
            except ValueError as error:
                self.fail(f"post-cutoff Treasury row blocked publication: {error}")

        self.assertEqual(bundle.detail.loc[0, "latest_date"], "2026-07-31")
        self.assertEqual(bundle.detail.loc[0, "latest_value"], 4.2)
        self.assertEqual(bundle.source_log.loc[0, "observations"], 2)

    def test_historical_world_bank_bundle_filters_future_month_end(self):
        config = next(
            row
            for row in load_macro_asset_universe()
            if row.series_code == "COMEX_GOLD"
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Monthly Prices"
        sheet.append(["Monthly Prices"])
        sheet.append(["Date", "Gold"])
        sheet.append(["", "$/troy oz"])
        sheet.append(["2026M04", 3200.0])
        sheet.append(["2026M05", 3300.0])
        sheet.append(["2026M06", 9999.0])
        stream = io.BytesIO()
        workbook.save(stream)
        workbook_bytes = stream.getvalue()
        page = (
            '<a href="https://thedocs.worldbank.org/official/'
            'CMO-Historical-Data-Monthly.xlsx">Monthly prices</a>'
        ).encode()

        def fake_get(_session, url, **_kwargs):
            body = workbook_bytes if url.endswith(".xlsx") else page
            return OfficialHttpResponse(
                body=body,
                url=url,
                headers={},
                trace=OfficialHttpTrace(1, 1, [200], url),
            )

        session = unittest.mock.Mock(headers={})
        with patch.object(
            macro_assets_module,
            "load_macro_asset_universe",
            return_value=[config],
        ), patch.object(
            macro_assets_module,
            "_session",
            return_value=session,
        ), patch.object(
            macro_assets_module,
            "official_get",
            side_effect=fake_get,
        ):
            try:
                bundle = fetch_macro_asset_bundle(as_of_date=date(2026, 6, 15))
            except ValueError as error:
                self.fail(f"future World Bank month-end blocked publication: {error}")

        self.assertEqual(bundle.detail.loc[0, "latest_date"], "2026-05-31")
        self.assertEqual(bundle.detail.loc[0, "latest_value"], 3300.0)
        self.assertEqual(
            bundle.commodity_price_history["observation_date"].tolist(),
            ["2026-04-30", "2026-05-31"],
        )

    def test_world_bank_discovery_rejects_non_world_bank_mirror(self):
        page = (
            '<a href="https://mirror.example/CMO-Historical-Data-Monthly.xlsx">'
            'Monthly prices</a>'
        )
        response = unittest.mock.Mock(
            content=page.encode(), text=page, status_code=200, headers={}
        )
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.return_value = response
        config = replace(
            self._config("world_bank_pink_sheet", "Gold"),
            source_url="https://www.worldbank.org/en/research/commodity-markets",
            level_unit="$/toz",
        )

        with self.assertRaisesRegex(ValueError, "official monthly workbook link"):
            _fetch_config_history(config, session)

        self.assertEqual(session.get.call_count, 1)

    def test_world_bank_discovery_rejects_misleading_monthly_link_labels(self):
        for label in ("Monthly historical archive", "Monthly price forecast"):
            with self.subTest(label=label):
                page = (
                    '<a href="https://thedocs.worldbank.org/official/'
                    f'CMO-Historical-Data-Monthly.xlsx">{label}</a>'
                )

                with self.assertRaisesRegex(
                    ValueError,
                    "official monthly workbook link",
                ):
                    _discover_world_bank_monthly_url(
                        page,
                        "https://www.worldbank.org/en/research/commodity-markets",
                    )

    def test_monthly_price_uses_latest_provider_bounded_month_end(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,"
                "source,source_url,frequency,level_unit,change_unit,sort_order,notes,"
                "freshness_days\n"
                "commodity,commodities,GOLD,黄金,Gold,world_bank_pink_sheet,Gold,"
                "World Bank,https://www.worldbank.org/en/research/commodity-markets,"
                "monthly,$/toz,pct,1,official monthly benchmark,45\n",
                encoding="utf-8",
            )
            history = [
                {"date": date(2026, 4, 30), "value": 3200.0, "unit": "$/toz"},
                {"date": date(2026, 5, 31), "value": 3300.0, "unit": "$/toz"},
                {"date": date(2026, 6, 30), "value": 9999.0, "unit": "$/toz"},
            ]

            def bounded_fetch(config, session, as_of_date=None):
                del config, session
                return (
                    [
                        point
                        for point in history
                        if point["date"] <= as_of_date
                    ],
                    b"workbook",
                    "https://thedocs.worldbank.org/monthly.xlsx",
                )

            with patch(
                "pipeline.internal.capital_weekly.macro_assets._fetch_config_history",
                side_effect=bounded_fetch,
            ):
                detail, source = fetch_macro_assets(
                    universe,
                    as_of_date=date(2026, 6, 15),
                )

        self.assertEqual(detail.loc[0, "latest_date"], "2026-05-31")
        self.assertEqual(detail.loc[0, "latest_value"], 3300.0)
        self.assertEqual(source.loc[0, "observations"], 2)

    def test_world_bank_price_staler_than_configured_45_days_blocks_publication(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,"
                "source,source_url,frequency,level_unit,change_unit,sort_order,notes,"
                "freshness_days\n"
                "commodity,commodities,GOLD,黄金,Gold,world_bank_pink_sheet,Gold,"
                "World Bank,https://www.worldbank.org/en/research/commodity-markets,"
                "monthly,$/toz,pct,1,official monthly benchmark,45\n",
                encoding="utf-8",
            )
            history = [
                {"date": date(2026, 5, 31), "value": 3200.0},
                {"date": date(2026, 6, 30), "value": 3300.0},
            ]
            with patch(
                "pipeline.internal.capital_weekly.macro_assets._fetch_config_history",
                return_value=(
                    history,
                    b"workbook",
                    "https://thedocs.worldbank.org/monthly.xlsx",
                ),
            ):
                detail, source = fetch_macro_assets(
                    universe,
                    as_of_date=date(2026, 8, 15),
                    allow_partial=True,
                )
                with self.assertRaisesRegex(ValueError, "GOLD"):
                    fetch_macro_assets(universe, as_of_date=date(2026, 8, 15))

        self.assertEqual(source.loc[0, "status"], "FETCH_FAILED")
        self.assertIn("stale beyond configured 45", source.loc[0, "error"])
        self.assertEqual(detail.loc[0, "qc_flag"], "FETCH_FAILED")

    def test_boe_shared_payload_uses_one_request_and_marks_cache_hit(self):
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        text = "DATE,IUDBEDR,IUDSOIA\n10 Jul 2026,4.25,3.97\n"
        response = unittest.mock.Mock(content=text.encode(), text=text)
        response.raise_for_status.return_value = None
        session.get.return_value = response
        _fetch_config_history(self._config("boe_iadb", "IUDBEDR"), session)
        _fetch_config_history(self._config("boe_iadb", "IUDSOIA"), session)
        self.assertEqual(session.get.call_count, 1)
        self.assertEqual(session._macro_attempt_trace[-1]["status"], "cache_hit")

    def test_hkma_and_boj_follow_official_pagination(self):
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        hkma_pages = [
            {"header": {"success": True, "total_count": 2}, "result": {"records": [{"end_of_date": "2026-07-09", "disc_win_base_rate": "4.75"}]}},
            {"header": {"success": True, "total_count": 2}, "result": {"records": [{"end_of_date": "2026-07-10", "disc_win_base_rate": "4.75"}]}},
        ]
        responses = []
        for payload in hkma_pages:
            text = json.dumps(payload); response = unittest.mock.Mock(content=text.encode(), text=text)
            response.raise_for_status.return_value = None; responses.append(response)
        session.get.side_effect = responses
        _fetch_config_history(self._config("hkma", "disc_win_base_rate"), session)
        self.assertEqual(session.get.call_count, 2)
        self.assertIn("pagesize=1000", session.get.call_args_list[0].args[0])
        self.assertIn("offset=1", session.get.call_args.args[0])

        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        pages = [
            {"RESULTSET": [{"SURVEY_DATES": ["2026-07-09"], "VALUES": ["0.01"]}], "NEXTPOSITION": 2},
            {"RESULTSET": [{"SURVEY_DATES": ["2026-07-10"], "VALUES": ["0.02"]}]},
        ]
        responses = []
        for payload in pages:
            text = json.dumps(payload); response = unittest.mock.Mock(content=text.encode(), text=text)
            response.raise_for_status.return_value = None; responses.append(response)
        session.get.side_effect = responses
        history, _, _ = _fetch_config_history(self._config("boj_api", "FM01:STRDCLUCON"), session)
        self.assertEqual([row["date"] for row in history], [date(2026, 7, 9), date(2026, 7, 10)])
        self.assertIn("startPosition=2", session.get.call_args.args[0])

    def test_hkma_falls_back_to_official_daily_pages_when_api_is_unavailable(self):
        def daily_page(day, value):
            text = (
                f'<meta name="date" content="{day.isoformat()}T00:00:00+08:00">'
                f'<p>Date and Time (日期和時間) : 18:30, '
                f'{day.strftime("%d/%m/%Y")}</p>'
                '<div>Base Rate (基本利率) :</div>'
                f'<div>{value:.2f}%</div>'
            )
            response = unittest.mock.Mock(
                content=text.encode("utf-8"),
                text=text,
            )
            response.raise_for_status.return_value = None
            return response

        session = unittest.mock.Mock(
            _macro_attempt_trace=[],
            _macro_raw_parts=[],
        )
        session.get.side_effect = [
            requests.ConnectionError("HKMA API unavailable"),
            daily_page(date(2026, 8, 21), 4.00),
            daily_page(date(2026, 8, 20), 4.00),
            daily_page(date(2026, 8, 14), 4.00),
            daily_page(date(2026, 7, 31), 4.25),
            daily_page(date(2025, 12, 31), 4.50),
        ]

        history, raw, provenance = _fetch_config_history(
            self._config("hkma", "disc_win_base_rate"),
            session,
            as_of_date=date(2026, 8, 21),
        )

        self.assertEqual(
            history,
            [
                {"date": date(2025, 12, 31), "value": 4.50},
                {"date": date(2026, 7, 31), "value": 4.25},
                {"date": date(2026, 8, 14), "value": 4.00},
                {"date": date(2026, 8, 20), "value": 4.00},
                {"date": date(2026, 8, 21), "value": 4.00},
            ],
        )
        self.assertIn(b"Base Rate", raw)
        self.assertIn("www.hkma.gov.hk", provenance)
        self.assertEqual(session.get.call_count, 6)

    def test_boj_repeated_next_position_fails_instead_of_looping(self):
        payload = {"RESULTSET": [{"SURVEY_DATES": ["2026-07-09"], "VALUES": ["0.01"]}], "NEXTPOSITION": 2}
        text = json.dumps(payload)
        response = unittest.mock.Mock(content=text.encode(), text=text)
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.return_value = response
        with self.assertRaisesRegex(ValueError, "repeated.*position"):
            _fetch_config_history(self._config("boj_api", "FM01:STRDCLUCON"), session)
        self.assertLessEqual(session.get.call_count, 2)

    def test_hkma_false_total_count_hits_page_guard(self):
        payload = {"header": {"success": True, "total_count": 999999}, "result": {"records": [{"end_of_date": "2026-07-09", "disc_win_base_rate": "4.75"}]}}
        text = json.dumps(payload)
        response = unittest.mock.Mock(content=text.encode(), text=text)
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.return_value = response
        with patch("pipeline.internal.capital_weekly.macro_assets.MAX_PROVIDER_PAGES", 3):
            with self.assertRaisesRegex(ValueError, "HKMA pagination exceeded"):
                _fetch_config_history(self._config("hkma", "disc_win_base_rate"), session)
        self.assertEqual(session.get.call_count, 3)

    def test_pboc_paginates_beyond_twelve_pages_until_start_is_covered(self):
        def response(text):
            item = unittest.mock.Mock(content=text.encode(), text=text)
            item.raise_for_status.return_value = None
            return item

        list_pages = []
        announcements = []
        for page in range(14):
            day = "2024年12月31日" if page == 13 else "2026年7月10日"
            href = f"/zhengcehuobisi/125207/125213/125431/125475/{1000 + page}/index.html"
            list_pages.append(response(f'<a href="{href}">{day}公开市场业务交易公告</a>'))
            announcements.append(response(f'<div>{day}</div><table><tr><td>期限</td><td>7天</td></tr><tr><td>操作利率</td><td>1.40%</td></tr></table>'))
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.side_effect = list_pages + announcements
        history, _, provenance = _fetch_config_history(self._config("pboc_omo", "7D_REVERSE_REPO"), session)
        self.assertTrue(history)
        self.assertIn("index13.html", provenance)
        self.assertEqual(session.get.call_count, 28)

    def test_pboc_page_cap_fails_if_requested_start_not_covered(self):
        text = '<a href="/zhengcehuobisi/125207/125213/125431/125475/1000/index.html">2026年7月10日公告</a>'
        response = unittest.mock.Mock(content=text.encode(), text=text)
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.return_value = response
        with patch("pipeline.internal.capital_weekly.macro_assets.MAX_PROVIDER_PAGES", 3):
            with self.assertRaisesRegex(ValueError, "PBOC OMO pagination exceeded 3 pages"):
                _fetch_config_history(self._config("pboc_omo", "7D_REVERSE_REPO"), session)
        self.assertEqual(session.get.call_count, 3)

    def test_pboc_empty_archive_page_is_not_start_coverage(self):
        response = unittest.mock.Mock(content=b"<html>empty</html>", text="<html>empty</html>")
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.return_value = response
        with self.assertRaisesRegex(ValueError, "before covering"):
            _fetch_config_history(self._config("pboc_omo", "7D_REVERSE_REPO"), session)

    def test_pboc_malformed_latest_announcement_fails_instead_of_carrying_old_rate(self):
        def response(text):
            item = unittest.mock.Mock(content=text.encode(), text=text)
            item.raise_for_status.return_value = None
            return item
        listing = response(
            '<a href="/zhengcehuobisi/125207/125213/125431/125475/20260710/index.html">2026年7月10日公告</a>'
            '<a href="/zhengcehuobisi/125207/125213/125431/125475/20241231/index.html">2024年12月31日公告</a>'
        )
        malformed_latest = response('<div>2026年7月10日</div><p>无法识别</p>')
        older_valid = response('<div>2024年12月31日</div><p>期限7天 操作利率1.50%</p>')
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.side_effect = [listing, malformed_latest, older_valid]
        with self.assertRaisesRegex(ValueError, "20260710/index"):
            _fetch_config_history(self._config("pboc_omo", "7D_REVERSE_REPO"), session)

    def test_pboc_mixed_valid_non_seven_day_and_seven_day_archive(self):
        def response(text):
            item = unittest.mock.Mock(content=text.encode(), text=text)
            item.raise_for_status.return_value = None
            return item
        listing = response(
            '<a href="/zhengcehuobisi/125207/125213/125431/125475/20260710/index.html">2026年7月10日公告</a>'
            '<a href="/zhengcehuobisi/125207/125213/125431/125475/20260709/index.html">2026年7月9日公告</a>'
            '<a href="/zhengcehuobisi/125207/125213/125431/125475/20241231/index.html">2024年12月31日公告</a>'
        )
        valid_fourteen_day = response('<div>2026年7月10日</div><p>期限14天 操作利率1.55%</p>')
        valid_seven_day = response('<div>2026年7月9日</div><p>期限7天 操作利率1.40%</p>')
        old_seven_day = response('<div>2024年12月31日</div><p>期限7天 操作利率1.50%</p>')
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.side_effect = [listing, valid_fourteen_day, valid_seven_day, old_seven_day]

        history, _, _ = _fetch_config_history(self._config("pboc_omo", "7D_REVERSE_REPO"), session)

        self.assertEqual(history[-1]["value"], 1.4)
        self.assertNotIn(1.55, {row["value"] for row in history})

    def test_pboc_follows_official_archive_next_links_and_parses_real_announcement_shape(self):
        def response(text):
            item = unittest.mock.Mock(content=text.encode(), text=text)
            item.raise_for_status.return_value = None
            return item

        first = response(
            '<a href="/zhengcehuobisi/125207/125213/125431/125475/2026071001/index.html">公告</a>'
            '<span class="hui12">2026-07-10</span>'
            '<a tagname="/zhengcehuobisi/125207/125213/125431/125475/17081-2.html">下一页</a>'
        )
        second = response(
            '<a href="/zhengcehuobisi/125207/125213/125431/125475/2024123101/index.html">公告</a>'
            '<span class="hui12">2024-12-31</span>'
        )
        announcement = response(
            '<meta name="Description" content="公开市场业务交易公告。2026年7月10日中国人民银行'
            '开展了200亿元7天期逆回购操作。具体情况如下。逆回购操作情况期限操作利率投标量'
            '中标量7天1.40%200亿元200亿元中国人民银行公开市场业务操作室">'
        )
        old_announcement = response(
            '<meta name="Description" content="公开市场业务交易公告。2024年12月31日中国人民银行'
            '开展了200亿元7天期逆回购操作。具体情况如下。期限操作利率7天1.50%">'
        )
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.side_effect = [first, second, announcement, old_announcement]
        history, _, provenance = _fetch_config_history(self._config("pboc_omo", "7D_REVERSE_REPO"), session)
        self.assertEqual(history[-1]["value"], 1.4)
        self.assertIn("17081-2.html", provenance)

    def test_chinamoney_fdr007_posts_dates_as_form_fields_and_parses_real_records_shape(self):
        text = json.dumps({"data": {"baseCurveCfgList": ["FDR007"]}, "records": [
            {"lfiProducDate": "2026-07-09", "frValueMap": {"date": "2026-07-09", "FDR007": "1.3900"}},
            {"lfiProducDate": "2026-07-10", "frValueMap": {"date": "2026-07-10", "FDR007": "1.3800"}},
        ]})
        response = unittest.mock.Mock(content=text.encode(), text=text)
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.post.return_value = response
        history, _, _ = _fetch_config_history(self._config("chinamoney_frr", "FDR007"), session)
        self.assertEqual([row["value"] for row in history], [1.39, 1.38])
        form = session.post.call_args.kwargs["data"]
        self.assertEqual(form["lang"], "CN")
        self.assertIn("startDate", form)
        self.assertIn("endDate", form)

    def test_blocked_real_provider_caches_raw_and_records_exact_provenance(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            raw_dir = Path(directory) / "raw"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
                "policy_rate,policy_rates,RBNZ_OCR,新西兰,OCR,rbnz_xlsx,Official Cash Rate (OCR),RBNZ,https://official.test/rbnz,daily,percent,bp,1,blocked\n",
                encoding="utf-8",
            )
            session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
            response = unittest.mock.Mock(content=b"official blocked body", text="official blocked body")
            response.raise_for_status.side_effect = RuntimeError("403 blocked")
            session.get.return_value = response
            with patch("pipeline.internal.capital_weekly.macro_assets._session", return_value=session):
                detail, source = fetch_macro_assets(
                    universe,
                    raw_dir,
                    allow_partial=True,
                )
            self.assertEqual(source.loc[0, "status"], "FETCH_FAILED")
            self.assertEqual(detail.loc[0, "qc_flag"], "FETCH_FAILED")
            self.assertIn("GET https://www.rbnz.govt.nz/", source.loc[0, "source_url"])
            self.assertIn("[attempting]", source.loc[0, "source_url"])
            self.assertEqual((raw_dir / "RBNZ_OCR.raw").read_bytes(), b"official blocked body")


    def test_policy_events_carry_to_business_as_of_but_market_history_stays_sparse(self):
        event = [{"date": date(2026, 7, 8), "value": 1.4}]
        market = [{"date": date(2026, 7, 8), "value": 1.41}]
        self.assertEqual(_carry_forward_business_daily(event, date(2026, 7, 10))[-1], {"date": date(2026, 7, 10), "value": 1.4})
        self.assertEqual(market, [{"date": date(2026, 7, 8), "value": 1.41}])

    def test_every_configured_provider_dispatches(self):
        contracts = {
            "us_treasury": ("GET", "home.treasury.gov"), "fred": ("GET", "fredgraph.csv"),
            "fred_millions_to_billions": ("GET", "fredgraph.csv"),
            "us_treasury_real": ("GET", "type=daily_treasury_real_yield_curve"),
            "yahoo_chart": ("GET", "query2.finance.yahoo.com"), "china_bond": ("POST", "yield.chinabond.com.cn"),
            "sina_fx": ("GET", "NewForexService.getDayKLine?symbol=fx_susdcnh"),
            "pboc_lpr": ("POST", "LprHisExcel"), "hkab_hibor": ("GET", "hkab.org.hk/api/hibor"),
            "nyfed_rates": ("GET", "markets.newyorkfed.org"), "ecb": ("GET", "data-api.ecb.europa.eu"),
            "boe_iadb": ("GET", "SeriesCodes=IUDBEDR,IUDSOIA"), "boj_policy": ("GET", "boj.or.jp/en/mopo/mpmdeci/index.htm"),
            "hkma": ("GET", "api.hkma.gov.hk"),
            "boc_valet": ("GET", "V39079,AVG.INTWO"),
            "snb_cube": ("GET", "data.snb.ch/api/cube"), "boj_api": ("GET", "stat-search.boj.or.jp"),
            "chinamoney_frr": ("POST", "FrrHis"),
            "eia_v2": ("GET", "api.eia.gov/v2"),
            "world_bank_pink_sheet": (
                "GET",
                "worldbank.org/en/research/commodity-markets",
            ),
        }
        universe = load_macro_asset_universe()
        self.assertEqual(set(contracts), {item.provider for item in universe} - {"calculated"})
        expected_symbols = {
            "us_treasury": {"2-year", "5-year", "10-year", "30-year"},
            "us_treasury_real": {"5-year", "10-year"},
            "fred": {"BAMLC0A0CM", "BAMLH0A0HYM2", "DFEDTARL", "DFEDTARU", "IORB", "RRPONTSYAWARD", "RRPONTSYD"},
            "fred_millions_to_billions": {"WALCL", "WTREGEN"},
            "yahoo_chart": {
                "HG=F", "DX-Y.NYB", "CNY=X",
                "HKD=X", "EURUSD=X", "JPY=X", "GBPUSD=X", "AUDUSD=X",
                "CAD=X", "CHF=X", "BTC-USD", "SPY", "TLT",
            }, "china_bond": {"2Y", "5Y", "10Y", "30Y"},
            "sina_fx": {"fx_susdcnh"},
            "pboc_lpr": {"1Y", "5Y+"}, "hkab_hibor": {"1M", "3M"},
            "nyfed_rates": {"EFFR", "SOFR"},
            "ecb": {"FM.D.U2.EUR.4F.KR.DFR.LEV", "FM.D.U2.EUR.4F.KR.MRR_FR.LEV", "FM.D.U2.EUR.4F.KR.MLFR.LEV", "EST.B.EU000A2X2A25.WT"},
            "boe_iadb": {"IUDBEDR", "IUDSOIA"}, "boj_policy": {"UNCOLLATERALIZED_ON_TARGET"},
            "hkma": {"disc_win_base_rate"},
            "boc_valet": {"V39079", "AVG.INTWO"},
            "snb_cube": {"snboffzisa:D0=LZ"}, "boj_api": {"FM01:STRDCLUCON"},
            "chinamoney_frr": {"FDR007"},
            "eia_v2": {"RWTC", "RBRTE", "RNGWHHD"},
            "world_bank_pink_sheet": {
                "Gold", "Copper", "Maize", "Soybeans", "Wheat, US SRW",
                "Rice, Thai 5%", "Cotton, A Index", "Sugar, world",
                "Coffee, Arabica", "Cocoa", "Beef **",
            },
        }
        self.assertEqual(
            expected_symbols,
            {provider: {item.provider_symbol for item in universe if item.provider == provider} for provider in contracts},
        )
        for provider, (method, endpoint) in contracts.items():
            config = next(item for item in load_macro_asset_universe() if item.provider == provider)
            session = unittest.mock.Mock()
            session._macro_attempt_trace = []
            session._macro_raw_parts = []
            response = unittest.mock.Mock(
                content=b"not a valid official payload",
                text="not a valid official payload",
                status_code=200,
                headers={},
            )
            response.raise_for_status.return_value = None
            session.get.return_value = response
            session.post.return_value = response
            with patch.dict(os.environ, {"EIA_API_KEY": "test-key"}):
                try:
                    _fetch_config_history(config, session)
                except Exception:
                    pass
            call = (session.get if method == "GET" else session.post).call_args_list[0]
            self.assertIn(endpoint, call.args[0], provider)
            expected_change_unit = (
                "pct"
                if provider
                in {
                    "yahoo_chart",
                    "sina_fx",
                    "eia_v2",
                    "world_bank_pink_sheet",
                }
                else "usd_billions"
                if provider == "fred_millions_to_billions"
                else "bp"
            )
            self.assertEqual(config.change_unit, expected_change_unit, provider)

    def test_configured_administered_boc_rate_carries_but_market_rate_does_not(self):
        text = json.dumps({"observations": [{"d": "2026-07-10", "V39079": {"v": "2.75"}, "AVG.INTWO": {"v": "2.70"}}]})
        response = unittest.mock.Mock(content=text.encode(), text=text)
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.return_value = response
        with patch("pipeline.internal.capital_weekly.macro_assets.date") as mocked_date:
            mocked_date.today.return_value = date(2026, 7, 13)
            target, _, _ = _fetch_config_history(self._config("boc_valet", "V39079"), session)
            market, _, _ = _fetch_config_history(self._config("boc_valet", "AVG.INTWO"), session)
        self.assertEqual(target[-1]["date"], date(2026, 7, 13))
        self.assertEqual(market[-1]["date"], date(2026, 7, 10))

    def test_lpr_and_hibor_reject_unknown_provider_symbols_before_http(self):
        session = unittest.mock.Mock()

        with self.assertRaisesRegex(ValueError, "Unsupported pboc_lpr symbol"):
            _fetch_config_history(self._config("pboc_lpr", "5Y-UNKNOWN"), session)
        with self.assertRaisesRegex(ValueError, "Unsupported hkab_hibor symbol"):
            _fetch_config_history(self._config("hkab_hibor", "6M"), session)

        session.get.assert_not_called()
        session.post.assert_not_called()

    def test_hibor_fetches_bounded_anchor_windows_with_all_snapshot_bases(self):
        config = self._config("hkab_hibor", "1M")
        second_config = self._config("hkab_hibor", "3M")
        session = unittest.mock.Mock()
        session._macro_attempt_trace = []
        session._macro_raw_parts = []
        values = {
            date(2025, 12, 31): 2.0,
            date(2026, 6, 30): 2.1,
            date(2026, 7, 3): 2.2,
            date(2026, 7, 9): 2.3,
            date(2026, 7, 10): 2.4,
        }

        def response_for(url, timeout):
            query = dict(part.split("=") for part in url.split("?", 1)[1].split("&"))
            day = date(int(query["year"]), int(query["month"]), int(query["day"]))
            payload = {"date": f"{day.year}-{day.month}-{day.day}", "isHoliday": day not in values}
            if day in values:
                payload["1 Month"] = values[day]
                payload["3 Months"] = values[day] + 0.1
            response = unittest.mock.Mock()
            response.content = json.dumps(payload).encode("utf-8")
            response.text = response.content.decode("utf-8")
            response.raise_for_status.return_value = None
            return response

        session.get.side_effect = response_for
        with patch("pipeline.internal.capital_weekly.macro_assets.date") as mocked_date:
            mocked_date.today.return_value = date(2026, 7, 12)
            history, _, _ = _fetch_config_history(config, session)
            second_history, _, _ = _fetch_config_history(second_config, session)

        snapshot = __import__("pipeline.internal.capital_weekly.returns", fromlist=["calculate_macro_snapshot"]).calculate_macro_snapshot(history, "bp")
        self.assertLessEqual(session.get.call_count, 56)
        self.assertEqual(snapshot.latest_date, date(2026, 7, 10))
        self.assertEqual(snapshot.daily_base_date, date(2026, 7, 9))
        self.assertEqual(snapshot.weekly_base_date, date(2026, 7, 3))
        self.assertEqual(snapshot.mtd_base_date, date(2026, 6, 30))
        self.assertEqual(snapshot.ytd_base_date, date(2025, 12, 31))
        self.assertEqual(snapshot.qc_flag, "OK")
        self.assertEqual([row["date"] for row in second_history], [row["date"] for row in history])

    def test_lpr_fixing_is_carried_forward_on_business_days_through_as_of_date(self):
        config = self._config("pboc_lpr", "1Y")
        session = unittest.mock.Mock()
        response = unittest.mock.Mock()
        response.content = b"PK fixture"
        response.raise_for_status.return_value = None

        with patch("pipeline.internal.capital_weekly.macro_assets.date") as mocked_date, patch(
            "pipeline.internal.capital_weekly.macro_assets._post", return_value=response
        ), patch(
            "pipeline.internal.capital_weekly.macro_assets._parse_lpr_xlsx",
            return_value=[
                {"date": date(2025, 12, 22), "value": 3.1},
                {"date": date(2026, 6, 22), "value": 3.0},
            ],
        ):
            mocked_date.today.return_value = date(2026, 7, 12)
            history, _, _ = _fetch_config_history(config, session)

        by_date = {row["date"]: row["value"] for row in history}
        self.assertEqual(by_date[date(2026, 7, 10)], 3.0)
        self.assertEqual(by_date[date(2026, 7, 9)], 3.0)
        self.assertNotIn(date(2026, 7, 11), by_date)
        snapshot = __import__("pipeline.internal.capital_weekly.returns", fromlist=["calculate_macro_snapshot"]).calculate_macro_snapshot(history, "bp")
        self.assertEqual(snapshot.latest_date, date(2026, 7, 10))
        self.assertEqual(snapshot.daily_base_date, date(2026, 7, 9))
        self.assertEqual(snapshot.daily_change, 0.0)

    def test_failure_audit_uses_attempt_trace_instead_of_generic_source_url(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
                "fixed_income,x,ONE,一,One,fred,ONE,FRED,https://generic.test,daily,percent,bp,1,note\n",
                encoding="utf-8",
            )

            def fail_with_trace(config, session):
                session._macro_attempt_trace = [
                    {"method": "POST", "url": "https://expanded.test/chunk-1", "status": "completed"},
                    {"method": "POST", "url": "https://expanded.test/chunk-2", "status": "attempting"},
                ]
                raise RuntimeError("chunk 2 failed")

            with patch("pipeline.internal.capital_weekly.macro_assets._fetch_config_history", side_effect=fail_with_trace):
                _, source_log = fetch_macro_assets(
                    universe,
                    allow_partial=True,
                )

        provenance = source_log.loc[0, "source_url"]
        self.assertIn("POST https://expanded.test/chunk-1 [completed]", provenance)
        self.assertIn("POST https://expanded.test/chunk-2 [attempting]", provenance)
        self.assertNotEqual(provenance, "https://generic.test")

    def test_eia_prepared_url_secret_is_redacted_from_every_audit_artifact(self):
        secret = "audit-sentinel-eia-key"
        prepared_url = (
            "https://api.eia.gov/v2/petroleum/pri/spt/data/"
            f"?api_key={secret}&frequency=daily"
        )
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            raw_dir = Path(directory) / "raw"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,"
                "provider_symbol,source,source_url,frequency,level_unit,"
                "change_unit,sort_order,notes,provider_route\n"
                "commodity,commodities,WTI,原油,WTI,eia_v2,RWTC,EIA,"
                "https://www.eia.gov/opendata/,daily,$/BBL,pct,1,note,"
                "petroleum/pri/spt\n",
                encoding="utf-8",
            )

            def fail_with_prepared_url(_config, session, as_of_date=None):
                del as_of_date
                session._macro_attempt_trace = [
                    {"method": "GET", "url": prepared_url, "status": "attempting"}
                ]
                session._macro_raw_parts = [prepared_url.encode("utf-8")]
                raise RuntimeError(f"401 Client Error for url: {prepared_url}")

            with patch.dict(os.environ, {"EIA_API_KEY": secret}), patch(
                "pipeline.internal.capital_weekly.macro_assets._fetch_config_history",
                side_effect=fail_with_prepared_url,
            ):
                detail, source_log = fetch_macro_assets(
                    universe,
                    raw_dir=raw_dir,
                    as_of_date=date(2026, 8, 9),
                    allow_partial=True,
                )
            serialized = "\n".join(
                (
                    detail.to_json(),
                    source_log.to_json(),
                    (raw_dir / "WTI.raw").read_text(encoding="utf-8"),
                )
            )

        self.assertNotIn(secret, serialized)
        self.assertIn("api_key=[REDACTED]", serialized)

    def test_raw_response_is_cached_when_parsing_fails(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            raw_dir = Path(directory) / "raw"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
                "fixed_income,x,ONE,一,One,fred,ONE,FRED,https://generic.test,daily,percent,bp,1,note\n",
                encoding="utf-8",
            )

            def parse_failure(config, session):
                session._macro_raw_parts = [b"upstream raw bytes"]
                session._macro_attempt_trace = [
                    {"method": "GET", "url": "https://expanded.test", "status": "completed"}
                ]
                raise ValueError("malformed CSV")

            with patch("pipeline.internal.capital_weekly.macro_assets._fetch_config_history", side_effect=parse_failure):
                _, source_log = fetch_macro_assets(
                    universe,
                    raw_dir=raw_dir,
                    allow_partial=True,
                )

            self.assertEqual((raw_dir / "ONE.raw").read_bytes(), b"upstream raw bytes")

        self.assertEqual(source_log.loc[0, "status"], "FETCH_FAILED")
        self.assertEqual(source_log.loc[0, "raw_cache_status"], "OK")
        self.assertEqual(source_log.loc[0, "raw_cache_error"], "")

    def test_raw_response_is_cached_when_snapshot_fails_after_fetch_returns(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            raw_dir = Path(directory) / "raw"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
                "fixed_income,x,ONE,一,One,fred,ONE,FRED,https://generic.test,daily,percent,bp,1,note\n",
                encoding="utf-8",
            )
            with patch(
                "pipeline.internal.capital_weekly.macro_assets._fetch_config_history",
                return_value=([{"date": date(2026, 7, 10), "value": 1}], b"valid upstream raw", "https://expanded.test"),
            ):
                _, source_log = fetch_macro_assets(
                    universe,
                    raw_dir=raw_dir,
                    allow_partial=True,
                )

            self.assertEqual((raw_dir / "ONE.raw").read_bytes(), b"valid upstream raw")

        self.assertEqual(source_log.loc[0, "status"], "FETCH_FAILED")
        self.assertEqual(source_log.loc[0, "raw_cache_status"], "OK")

    def test_atomic_raw_cache_stages_in_destination_and_replaces_without_temp_leftover(self):
        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "raw" / "series.json"
            real_replace = os.replace
            calls = []

            def recording_replace(source, target):
                calls.append((Path(source), Path(target)))
                real_replace(source, target)

            with patch("pipeline.internal.capital_weekly.macro_assets.os.replace", side_effect=recording_replace):
                _atomic_write_bytes(destination, b"raw fixture")

            self.assertEqual(destination.read_bytes(), b"raw fixture")
            self.assertEqual(calls[0][0].parent, destination.parent)
            self.assertEqual(calls[0][1], destination)
            self.assertEqual(list(destination.parent.glob("*.tmp")), [])

    def test_raw_cache_failure_is_a_separate_audit_failure(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
                "commodity,x,ONE,一,One,yahoo_chart,X=F,Yahoo,https://example.test,daily,usd,pct,1,note\n",
                encoding="utf-8",
            )
            history = [
                {"date": date(2025, 12, 31), "value": 10},
                {"date": date(2026, 7, 9), "value": 11},
                {"date": date(2026, 7, 10), "value": 12},
            ]
            with patch("pipeline.internal.capital_weekly.macro_assets._fetch_config_history", return_value=(history, b"raw", "https://expanded.test")), patch(
                "pipeline.internal.capital_weekly.macro_assets._atomic_write_bytes", side_effect=OSError("disk full")
            ):
                detail, source_log = fetch_macro_assets(universe, raw_dir=Path(directory) / "raw")

        self.assertEqual(source_log.loc[0, "status"], "OK")
        self.assertEqual(source_log.loc[0, "raw_cache_status"], "CACHE_WRITE_FAILED")
        self.assertEqual(source_log.loc[0, "raw_cache_error"], "disk full")
        self.assertNotEqual(detail.loc[0, "qc_flag"], "FETCH_FAILED")

    def test_required_fetch_failure_blocks_partial_bundle_but_diagnostics_remain_available(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            header = (
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,"
                "source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
            )
            universe.write_text(
                header
                + "fixed_income,x,GOOD,好,Good,fred,GOOD,FRED,https://example.test,daily,percent,bp,1,ok\n"
                + "fixed_income,x,BAD,坏,Bad,fred,BAD,FRED,https://example.test,daily,percent,bp,2,bad\n",
                encoding="utf-8",
            )
            history = [
                {"date": date(2025, 12, 31), "value": 1.0},
                {"date": date(2026, 7, 9), "value": 1.1},
                {"date": date(2026, 7, 10), "value": 1.2},
            ]

            def fake_fetch(config, session):
                if config.series_code == "BAD":
                    raise RuntimeError("upstream unavailable")
                return history, b"fixture", "https://expanded.test/GOOD"

            with patch("pipeline.internal.capital_weekly.macro_assets._fetch_config_history", side_effect=fake_fetch):
                with self.assertRaisesRegex(ValueError, "BAD"):
                    fetch_macro_assets(universe)
                detail, source_log = fetch_macro_assets(
                    universe,
                    allow_partial=True,
                )

        self.assertEqual(detail["series_code"].tolist(), ["GOOD", "BAD"])
        self.assertEqual(source_log["series_code"].tolist(), ["GOOD", "BAD"])
        self.assertEqual(source_log["status"].tolist(), ["OK", "FETCH_FAILED"])
        self.assertEqual(detail.loc[1, "qc_flag"], "FETCH_FAILED")
        self.assertIn("upstream unavailable", source_log.loc[1, "error"])

    def test_optional_public_proxy_failure_does_not_block_bundle(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            universe.write_text(
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,"
                "source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
                "cross_asset,cross_asset_correlation,OPTIONAL_PROXY,可选代理,"
                "Optional proxy,yahoo_chart,OPTIONAL,Public proxy,"
                "https://example.test/proxy,daily,index,pct,1,optional\n",
                encoding="utf-8",
            )
            with patch(
                "pipeline.internal.capital_weekly.macro_assets._fetch_config_history",
                side_effect=RuntimeError("optional upstream unavailable"),
            ):
                try:
                    bundle = fetch_macro_asset_bundle(universe)
                except ValueError as error:
                    self.fail(f"optional public proxy blocked publication: {error}")

        self.assertEqual(bundle.detail.loc[0, "qc_flag"], "FETCH_FAILED")
        self.assertEqual(bundle.source_log.loc[0, "status"], "FETCH_FAILED")
        self.assertEqual(bundle.source_log.loc[0, "requiredness"], "optional")
        self.assertIn("optional upstream unavailable", bundle.source_log.loc[0, "error"])

    def test_calculated_series_emit_formula_and_provider_audit_metadata(self):
        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            header = (
                "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,"
                "source,source_url,frequency,level_unit,change_unit,sort_order,notes,"
                "calculation_id,formula_version,input_series_codes\n"
            )
            universe.write_text(
                header
                + "fixed_income,sovereign_curve,UST5Y,名义5年,Nominal 5Y,"
                "us_treasury,5-year,Treasury,https://example.test/nominal,daily,"
                "percent,bp,1,,,,\n"
                + "fixed_income,sovereign_curve,UST_REAL5Y,实际5年,Real 5Y,"
                "us_treasury_real,5-year,Treasury,https://example.test/real,daily,"
                "percent,bp,2,,,,\n"
                + "fixed_income,inflation_expectations,US_BE5Y,5年盈亏平衡,5Y Breakeven,"
                "calculated,UST5Y-UST_REAL5Y,Calculated,https://example.test/calculated,"
                "daily,percent,bp,3,Registered calculation,breakeven,breakeven-v1,"
                "UST5Y|UST_REAL5Y\n",
                encoding="utf-8",
            )

            def fake_fetch(config, session, as_of_date=None):
                value = 4.0 if config.series_code == "UST5Y" else 1.9
                return (
                    [
                        {"date": date(2025, 12, 31), "value": value - 0.5},
                        {"date": date(2026, 8, 7), "value": value},
                    ],
                    b"fixture",
                    config.source_url,
                )

            with patch(
                "pipeline.internal.capital_weekly.macro_assets._fetch_config_history",
                side_effect=fake_fetch,
            ):
                detail, source_log = fetch_macro_assets(
                    universe,
                    as_of_date=date(2026, 8, 9),
                )

        calculated = detail.loc[detail["series_code"] == "US_BE5Y"].iloc[0]
        self.assertAlmostEqual(calculated["latest_value"], 2.1)
        self.assertEqual(calculated["calculation_id"], "breakeven")
        self.assertEqual(calculated["formula_version"], "breakeven-v1")
        self.assertEqual(calculated["input_series_codes"], "UST5Y|UST_REAL5Y")
        calculated_audit = source_log.loc[
            source_log["series_code"] == "US_BE5Y"
        ].iloc[0]
        self.assertEqual(calculated_audit["provider"], "calculated")
        self.assertEqual(calculated_audit["source_tier"], "official")
        self.assertEqual(calculated_audit["requiredness"], "required")
        self.assertEqual(calculated_audit["provider_version"], "1.0.0")
        self.assertEqual(calculated_audit["schema_version"], "macro-asset-v3")
        self.assertEqual(calculated_audit["known_as_of"], "2026-08-07")
        self.assertEqual(calculated_audit["calculation_id"], "breakeven")
        self.assertEqual(calculated_audit["formula_version"], "breakeven-v1")
        self.assertEqual(
            calculated_audit["input_series_codes"],
            "UST5Y|UST_REAL5Y",
        )

    def test_treasury_csv_parser_normalizes_dates_values_and_blanks(self):
        fixture = "Date,2 Yr,10 Yr\n07/10/2026,4.21,4.56\n07/09/2026,,4.50\n"

        result = _parse_treasury_csv(fixture, "2 Yr")

        self.assertEqual(result, [{"date": date(2026, 7, 10), "value": 4.21}])

    def test_treasury_real_provider_uses_official_real_curve_fields_and_type(self):
        def response_for(url, timeout):
            fixture = (
                "Date,5 YR,10 YR\n"
                if "/2025/" in url
                else "Date,5 YR,10 YR\n08/07/2026,1.90,2.00\n"
            )
            response = unittest.mock.Mock(content=fixture.encode(), text=fixture)
            response.raise_for_status.return_value = None
            return response

        for symbol, expected_value in (("5-year", 1.9), ("10-year", 2.0)):
            with self.subTest(symbol=symbol):
                session = unittest.mock.Mock(
                    _macro_attempt_trace=[],
                    _macro_raw_parts=[],
                )
                session.get.side_effect = response_for

                history, _, provenance = _fetch_config_history(
                    self._config("us_treasury_real", symbol),
                    session,
                    as_of_date=date(2026, 8, 9),
                )

                self.assertEqual(
                    history,
                    [{"date": date(2026, 8, 7), "value": expected_value}],
                )
                self.assertEqual(session.get.call_count, 2)
                for call in session.get.call_args_list:
                    self.assertIn(
                        "type=daily_treasury_real_yield_curve",
                        call.args[0],
                    )
                self.assertIn(
                    "type=daily_treasury_real_yield_curve",
                    provenance,
                )

    def test_fred_csv_parser_drops_dot_missing_values(self):
        fixture = "observation_date,BAMLC0A0CM\n2026-07-08,.\n2026-07-09,0.76\n"

        result = _parse_fred_csv(fixture, "BAMLC0A0CM")

        self.assertEqual(result, [{"date": date(2026, 7, 9), "value": 0.76}])

    def test_h41_fred_provider_normalizes_millions_to_billions(self):
        config = self._config("fred_millions_to_billions", "WALCL")
        response = unittest.mock.Mock(
            content=b"observation_date,WALCL\n2026-08-05,8100000\n",
            text="observation_date,WALCL\n2026-08-05,8100000\n",
        )
        response.raise_for_status.return_value = None
        session = unittest.mock.Mock(_macro_attempt_trace=[], _macro_raw_parts=[])
        session.get.return_value = response

        history, _, _ = _fetch_config_history(
            config,
            session,
            as_of_date=date(2026, 8, 9),
        )

        self.assertEqual(
            history,
            [{"date": date(2026, 8, 5), "value": 8100.0}],
        )

    def test_yahoo_chart_parser_pairs_timestamps_and_closes_and_drops_nulls(self):
        fixture = json.dumps({"chart": {"result": [{
            "timestamp": [1783641600, 1783728000],
            "indicators": {"quote": [{"close": [100.5, None]}]},
        }], "error": None}})

        result = _parse_yahoo_chart(fixture)

        self.assertEqual(result, [{"date": date(2026, 7, 10), "value": 100.5}])

    def test_universe_has_report_required_fx_and_btc_with_unique_pct_codes(self):
        universe = load_macro_asset_universe()
        required = {"DXY", "USD_CNY", "USD_CNH", "USD_HKD", "BTC_USD"}
        configured = {item.series_code: item for item in universe if item.series_code in required}

        self.assertEqual(set(configured), required)
        self.assertEqual(len({item.series_code for item in universe}), len(universe))
        self.assertTrue(all(item.change_unit == "pct" for item in configured.values()))
        self.assertEqual(configured["BTC_USD"].asset_class, "commodity")
        self.assertTrue(all(
            configured[code].asset_class == "foreign_exchange"
            for code in required - {"BTC_USD"}
        ))

    def test_commodity_taxonomy_is_additive_to_existing_macro_configs(self):
        universe = load_macro_asset_universe()
        wti = next(row for row in universe if row.series_code == "WTI")

        self.assertEqual(wti.commodity_code, "WTI")
        self.assertEqual(wti.commodity_family, "refined_products")
        self.assertEqual(wti.price_kind, "official_cash")
        self.assertEqual(wti.provider, "eia_v2")
        self.assertEqual(
            next(row for row in universe if row.series_code == "UST2Y").commodity_code,
            "",
        )

    def test_usd_pairs_retain_report_direction_and_yahoo_history_discards_nulls(self):
        universe = {item.series_code: item for item in load_macro_asset_universe()}

        self.assertEqual(universe["USD_CNY"].provider_symbol, "CNY=X")
        self.assertEqual(universe["USD_CNH"].provider, "sina_fx")
        self.assertEqual(universe["USD_CNH"].provider_symbol, "fx_susdcnh")
        self.assertEqual(universe["USD_CNH"].source, "Sina Finance FX")
        self.assertEqual(universe["USD_HKD"].provider_symbol, "HKD=X")
        fixture = json.dumps({"chart": {"result": [{
            "timestamp": [1783641600, 1783728000],
            "indicators": {"quote": [{"close": [7.25, None]}]},
        }], "error": None}})
        self.assertEqual(
            _parse_yahoo_chart(fixture),
            [{"date": date(2026, 7, 10), "value": 7.25}],
        )

    def test_curve_spread_inner_joins_dates_and_subtracts_two_year_from_ten_year(self):
        ten_year = [
            {"date": "2026-07-01", "value": 4.40},
            {"date": "2026-07-02", "value": 4.35},
            {"date": "2026-07-04", "value": 4.30},
        ]
        two_year = [
            {"date": "2026-07-01", "value": 3.90},
            {"date": "2026-07-03", "value": 3.85},
            {"date": "2026-07-04", "value": 3.80},
        ]

        result = align_curve_spread(ten_year, two_year)

        self.assertEqual(
            [point["date"] for point in result],
            [date(2026, 7, 1), date(2026, 7, 4)],
        )
        self.assertAlmostEqual(result[0]["value"], 0.5)
        self.assertAlmostEqual(result[1]["value"], 0.5)

    def test_five_year_five_year_uses_registered_compounding_formula(self):
        expected = (
            ((1.0 + 2.4 / 100.0) ** 2) / (1.0 + 2.1 / 100.0) - 1.0
        ) * 100.0

        self.assertAlmostEqual(
            calculate_five_year_five_year(2.1, 2.4),
            expected,
            places=12,
        )

    def test_five_year_five_year_rejects_invalid_breakeven_inputs(self):
        for be5, be10 in (
            (-100.0, 2.4),
            (2.1, -100.0),
            (float("nan"), 2.4),
        ):
            with self.subTest(be5=be5, be10=be10):
                with self.assertRaisesRegex(
                    ValueError,
                    "Breakeven inputs",
                ):
                    calculate_five_year_five_year(be5, be10)

    def test_breakeven_uses_only_dates_shared_by_nominal_and_real_curves(self):
        result = align_series_histories(
            {
                "UST5Y": [{"date": date(2026, 8, 7), "value": 4.0}],
                "UST_REAL5Y": [
                    {"date": date(2026, 8, 6), "value": 1.8},
                    {"date": date(2026, 8, 7), "value": 1.9},
                ],
            },
            ("UST5Y", "UST_REAL5Y"),
            lambda nominal, real: nominal - real,
        )

        self.assertEqual(
            result,
            [{"date": date(2026, 8, 7), "value": 2.1}],
        )

    def test_universe_has_approved_81_series_and_registered_calculations(self):
        universe = load_macro_asset_universe()

        self.assertEqual(len(universe), 81)
        self.assertEqual(
            len({item.series_code for item in universe}),
            len(universe),
        )
        by_class = {
            asset_class: [item for item in universe if item.asset_class == asset_class]
            for asset_class in (
                "fixed_income",
                "commodity",
                "foreign_exchange",
                "policy_rate",
                "money_market",
                "liquidity",
                "calculation_input",
                "cross_asset",
            )
        }
        self.assertEqual(
            {asset_class: len(items) for asset_class, items in by_class.items()},
            {
                "fixed_income": 22,
                "commodity": 16,
                "foreign_exchange": 10,
                "policy_rate": 12,
                "money_market": 7,
                "liquidity": 4,
                "calculation_input": 2,
                "cross_asset": 8,
            },
        )
        new_rates = {
            item.series_code: item
            for item in universe
            if item.series_code in {
                "UST5Y",
                "UST_REAL5Y",
                "UST_REAL10Y",
                "US_BE5Y",
                "US_BE10Y",
                "US_5Y5Y",
            }
        }
        self.assertEqual(
            set(new_rates),
            {
                "UST5Y",
                "UST_REAL5Y",
                "UST_REAL10Y",
                "US_BE5Y",
                "US_BE10Y",
                "US_5Y5Y",
            },
        )
        self.assertTrue(
            all(item.change_unit == "bp" for item in new_rates.values())
        )
        self.assertEqual(new_rates["US_BE5Y"].calculation_id, "breakeven")
        self.assertEqual(new_rates["US_BE5Y"].formula_version, "breakeven-v1")
        self.assertEqual(
            new_rates["US_BE5Y"].input_series_codes,
            "UST5Y|UST_REAL5Y",
        )
        self.assertEqual(
            new_rates["US_5Y5Y"].formula_version,
            "forward-inflation-v1",
        )
        self.assertEqual(
            new_rates["US_5Y5Y"].input_series_codes,
            "US_BE5Y|US_BE10Y",
        )
        additions = by_class["policy_rate"] + by_class["money_market"]
        self.assertTrue(all(item.level_unit == "percent" for item in additions))
        self.assertTrue(all(item.change_unit == "bp" for item in additions))
        self._assert_policy_universe_semantics(universe)
        self.assertEqual(sum(item.series_code.startswith("LPR") for item in universe), 2)
        self.assertEqual(sum(item.series_code.startswith("HIBOR") for item in universe), 2)
        self.assertNotIn("PBOC_7D_RR", {item.series_code for item in universe})
        wave_1_codes = {
            "UST30Y5Y", "USHY_IG_OAS", "FED_TOTAL_ASSETS", "TGA_BALANCE",
            "ON_RRP_TAKE_UP", "FED_NET_LIQUIDITY", "COMEX_COPPER", "EUR_USD",
            "USD_JPY", "GBP_USD", "AUD_USD", "USD_CAD", "USD_CHF",
            "SPY_CLOSE_PROXY", "TLT_CLOSE_PROXY",
            "US_STOCK_BOND_CORR_13W", "US_STOCK_BOND_CORR_26W",
            "EQUITY_USD_CORR_13W", "EQUITY_USD_CORR_26W",
            "GOLD_REAL_YIELD_CORR_13W", "GOLD_REAL_YIELD_CORR_26W",
            "OIL_BREAKEVEN_CORR_13W", "OIL_BREAKEVEN_CORR_26W",
        }
        self.assertEqual(
            {item.series_code for item in universe}.intersection(wave_1_codes),
            wave_1_codes,
        )
        configured = {item.series_code: item for item in universe}
        self.assertEqual(configured["ON_RRP_TAKE_UP"].provider_symbol, "RRPONTSYD")
        self.assertEqual(
            configured["FED_NET_LIQUIDITY"].input_series_codes,
            "FED_TOTAL_ASSETS|TGA_BALANCE|ON_RRP_TAKE_UP",
        )
        self.assertEqual(configured["UST30Y5Y"].input_series_codes, "UST30Y|UST5Y")
        self.assertEqual(
            configured["USHY_IG_OAS"].input_series_codes,
            "USHY_OAS|USIG_OAS",
        )
        self.assertEqual(
            (
                macro_assets.CORRELATION_SPECS[
                    "US_STOCK_BOND_CORR_13W"
                ].window,
                macro_assets.CORRELATION_SPECS[
                    "US_STOCK_BOND_CORR_13W"
                ].minimum_observations,
            ),
            (65, 52),
        )
        self.assertEqual(
            (
                macro_assets.CORRELATION_SPECS[
                    "GOLD_REAL_YIELD_CORR_13W"
                ].left_transform,
                macro_assets.CORRELATION_SPECS[
                    "GOLD_REAL_YIELD_CORR_13W"
                ].right_transform,
            ),
            ("pct_return", "level_change"),
        )

    def test_merged_gold_inputs_produce_both_daily_proxy_correlations(self):
        configured = {
            row.series_code: row
            for row in load_macro_asset_universe()
        }
        universe = [
            configured["UST_REAL10Y"],
            configured["COMEX_GOLD"],
            configured["GOLD_REAL_YIELD_CORR_13W"],
            configured["GOLD_REAL_YIELD_CORR_26W"],
        ]
        business_days = [
            timestamp.date()
            for timestamp in pd.bdate_range("2026-01-05", "2026-08-28")
        ]
        treasury = "Date,10 YR\n" + "".join(
            f"{day.strftime('%m/%d/%Y')},{1.5 + index * 0.002 + (index % 11) * 0.003}\n"
            for index, day in enumerate(business_days)
        )
        yahoo = json.dumps(
            {
                "chart": {
                    "result": [
                        {
                            "timestamp": [
                                int(pd.Timestamp(day, tz="UTC").timestamp())
                                for day in business_days
                            ],
                            "indicators": {
                                "quote": [
                                    {
                                        "close": [
                                            2000.0
                                            + index * 1.75
                                            + (index % 7) * 2.5
                                            for index, _day in enumerate(business_days)
                                        ]
                                    }
                                ]
                            },
                        }
                    ],
                    "error": None,
                }
            }
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Monthly Prices"
        sheet.append(["Monthly Prices"])
        sheet.append(["Date", "Gold"])
        sheet.append(["", "$/troy oz"])
        for month, value in (
            ("2026M04", 3200.0),
            ("2026M05", 3250.0),
            ("2026M06", 3350.0),
            ("2026M07", 3400.0),
        ):
            sheet.append([month, value])
        stream = io.BytesIO()
        workbook.save(stream)
        workbook_bytes = stream.getvalue()
        page = (
            '<a href="https://thedocs.worldbank.org/official/'
            'CMO-Historical-Data-Monthly.xlsx">Monthly prices</a>'
        ).encode()

        def public_get(url, **_kwargs):
            if "/2025/" in url:
                text = "Date,10 YR\n"
            elif "home.treasury.gov" in url:
                text = treasury
            elif "query2.finance.yahoo.com" in url:
                text = yahoo
            else:
                raise AssertionError(f"unexpected public URL: {url}")
            response = unittest.mock.Mock(content=text.encode(), text=text)
            response.raise_for_status.return_value = None
            return response

        def official_get_fixture(_session, url, **_kwargs):
            body = workbook_bytes if url.endswith(".xlsx") else page
            return OfficialHttpResponse(
                body=body,
                url=url,
                headers={},
                trace=OfficialHttpTrace(1, 1, [200], url),
            )

        session = unittest.mock.Mock(headers={})
        session.get.side_effect = public_get
        with patch.object(
            macro_assets_module,
            "load_macro_asset_universe",
            return_value=universe,
        ), patch.object(
            macro_assets_module,
            "_session",
            return_value=session,
        ), patch.object(
            macro_assets_module,
            "official_get",
            side_effect=official_get_fixture,
        ):
            bundle = fetch_macro_asset_bundle(as_of_date=date(2026, 8, 30))

        detail = bundle.detail.set_index("series_code")
        source_log = bundle.source_log.set_index("series_code")
        self.assertEqual(detail.loc["COMEX_GOLD", "provider"], "world_bank_pink_sheet")
        self.assertEqual(detail.loc["COMEX_GOLD", "latest_value"], 3400.0)
        self.assertEqual(
            set(bundle.commodity_price_history["series_code"]),
            {"COMEX_GOLD"},
        )
        gold_history = bundle.commodity_price_history.sort_values(
            "observation_date"
        ).reset_index(drop=True)
        self.assertEqual(gold_history.loc[3, "value"], 3400.0)
        self.assertEqual(
            set(gold_history["source"]),
            {"World Bank Commodity Price Data (Pink Sheet)"},
        )
        self.assertEqual(
            set(gold_history["source_url"]),
            {
                "https://thedocs.worldbank.org/official/"
                "CMO-Historical-Data-Monthly.xlsx"
            },
        )
        self.assertFalse(
            gold_history["source_url"].str.contains(
                "query2.finance.yahoo.com",
                regex=False,
            ).any()
        )
        for series_code, observations in (
            ("GOLD_REAL_YIELD_CORR_13W", 65),
            ("GOLD_REAL_YIELD_CORR_26W", 130),
        ):
            with self.subTest(series_code=series_code):
                self.assertNotEqual(
                    detail.loc[series_code, "qc_flag"],
                    "FETCH_FAILED",
                    source_log.loc[series_code, "error"],
                )
                self.assertIsNotNone(detail.loc[series_code, "latest_value"])
                self.assertEqual(
                    detail.loc[series_code, "correlation_observations"],
                    observations,
                )
                self.assertIn(
                    "query2.finance.yahoo.com",
                    source_log.loc[series_code, "source_url"],
                )

    def test_correlation_proxy_rejects_a_non_https_source_url(self):
        config = next(
            row
            for row in load_macro_asset_universe()
            if row.series_code == "COMEX_GOLD"
        )

        with self.assertRaisesRegex(
            ValueError,
            "correlation proxy source_url must be a valid HTTPS URL",
        ):
            macro_assets_module._correlation_proxy_config(
                replace(
                    config,
                    correlation_proxy_source_url=(
                        "http://insecure.example/GC=F"
                    ),
                )
            )

    def test_failed_detail_omits_internal_source_description_like_success(self):
        good = replace(
            self._config("fred", "GOOD"),
            series_code="GOOD",
            source_description="success parser identity",
        )
        bad = replace(
            self._config("fred", "BAD"),
            series_code="BAD",
            sort_order=2,
            source_description="failure parser identity",
        )

        def fetch(config, _session, as_of_date=None):
            del as_of_date
            if config.series_code == "BAD":
                raise RuntimeError("fixture failure")
            return (
                [
                    {"date": date(2026, 8, 7), "value": 1.0},
                    {"date": date(2026, 8, 14), "value": 1.1},
                ],
                b"fixture",
                config.source_url,
            )

        with patch.object(
            macro_assets_module,
            "load_macro_asset_universe",
            return_value=[good, bad],
        ), patch.object(
            macro_assets_module,
            "_fetch_config_history",
            side_effect=fetch,
        ):
            bundle = fetch_macro_asset_bundle(
                as_of_date=date(2026, 8, 16),
                allow_partial=True,
            )

        self.assertNotIn("source_description", bundle.detail.columns)
        self.assertNotIn("freshness_days", bundle.detail.columns)

    def test_macro_source_audit_distinguishes_official_and_optional_proxy_rows(self):
        header = (
            "asset_class,group,series_code,name_cn,name_en,provider,provider_symbol,"
            "source,source_url,frequency,level_unit,change_unit,sort_order,notes\n"
        )
        rows = (
            "fixed_income,credit_spreads,OFFICIAL,官方,Official,fred,OFFICIAL,FRED,"
            "https://example.test/official,daily,percent,bp,1,Official row\n"
            "foreign_exchange,foreign_exchange,PROXY,代理,Proxy,yahoo_chart,PROXY=X,"
            "Yahoo public proxy,https://example.test/proxy,daily,index,pct,2,Proxy row\n"
        )
        history = [
            {"date": date(2025, 12, 31), "value": 100.0},
            {"date": date(2026, 8, 7), "value": 101.0},
        ]

        with tempfile.TemporaryDirectory() as directory:
            universe = Path(directory) / "universe.csv"
            universe.write_text(header + rows, encoding="utf-8")
            with patch(
                "pipeline.internal.capital_weekly.macro_assets._fetch_config_history",
                return_value=(history, b"fixture", "https://example.test/value"),
            ):
                _, source_log = fetch_macro_assets(
                    universe,
                    as_of_date=date(2026, 8, 9),
                )

        audit = source_log.set_index("series_code")
        self.assertEqual(
            (
                audit.loc["OFFICIAL", "source_tier"],
                audit.loc["OFFICIAL", "requiredness"],
            ),
            ("official", "required"),
        )
        self.assertEqual(
            (
                audit.loc["PROXY", "source_tier"],
                audit.loc["PROXY", "requiredness"],
            ),
            ("public_proxy", "optional"),
        )

    def test_policy_semantics_reject_persistent_china_mlf_aliases_but_allow_ecb_mlf(self):
        universe = load_macro_asset_universe()
        bad_universe = [*universe, replace(universe[0], series_code="PBOC_MLF", name_en="China MLF Rate")]

        self._assert_policy_universe_semantics(universe)
        with self.assertRaises(AssertionError):
            self._assert_policy_universe_semantics(bad_universe)

    def test_policy_semantics_rejects_standalone_dr007_even_on_fdr007_row(self):
        universe = load_macro_asset_universe()
        bad_universe = [
            replace(item, provider_symbol="DR007")
            if item.series_code == "CNY_FDR007" else item
            for item in universe
        ]

        with self.assertRaises(AssertionError):
            self._assert_policy_universe_semantics(bad_universe)

    def test_cli_writes_strict_json_with_null_for_failed_numeric_values(self):
        from pipeline.internal.scripts import fetch_macro_assets as fetch_cli

        detail = pd.DataFrame({
            "asset_class": ["fixed_income", "commodity"],
            "group": ["sovereign_curve", "commodities"],
            "series_code": ["UST10Y", "GOLD"],
            "name_cn": ["美国10年期国债收益率", "黄金"],
            "level_unit": ["percent", "usd_per_ounce"],
            "change_unit": ["bp", "pct"],
            "sort_order": [1, 2],
            "daily_change": [1.0, float("nan")],
            "weekly_change": [2.0, float("nan")],
            "mtd_change": [3.0, float("nan")],
            "ytd_change": [4.0, float("nan")],
            "qc_flag": ["OK", "FETCH_FAILED"],
        })
        source_log = pd.DataFrame({
            "series_code": ["UST10Y", "GOLD"],
            "status": ["OK", "FETCH_FAILED"],
        })

        with tempfile.TemporaryDirectory() as directory, patch(
            "pipeline.internal.scripts.fetch_macro_assets.fetch_macro_assets",
            return_value=(detail, source_log),
        ), patch.object(
            os.sys, "argv", ["fetch_macro_assets.py", "--output-dir", directory]
        ):
            fetch_cli.main()

            root = Path(directory)
            self.assertTrue((root / "fixed_income.csv").exists())
            self.assertTrue((root / "commodities.csv").exists())
            self.assertTrue((root / "commodity_price_history.csv").exists())
            self.assertEqual(
                (root / "commodity_price_history.csv")
                .read_text(encoding="utf-8")
                .splitlines()[0]
                .split(","),
                list(macro_assets_module.PRICE_HISTORY_FIELDS),
            )
            self.assertTrue((root / "foreign_exchange.csv").exists())
            self.assertTrue((root / "macro_divergence.csv").exists())
            self.assertTrue((root / "source_log.csv").exists())
            self.assertTrue((root / "raw").is_dir())
            snapshot = json.loads(
                (root / "macro_assets_snapshot.json").read_text(encoding="utf-8"),
                parse_constant=lambda value: (_ for _ in ()).throw(
                    ValueError(f"non-finite constant: {value}")
                ),
            )

        self.assertIsNone(snapshot["commodities"][0]["weekly_change"])
        self.assertEqual(snapshot["commodity_price_history"], [])

    def test_cli_publishes_all_five_asset_classes_and_full_snapshot_contract(self):
        from pipeline.internal.scripts import fetch_macro_assets as fetch_cli

        class_counts = {
            "fixed_income": 20,
            "commodity": 15,
            "foreign_exchange": 4,
            "policy_rate": 12,
            "money_market": 7,
        }
        groups = {
            "fixed_income": ("sovereign_curve", "policy_money_market", "credit_spreads"),
            "commodity": ("commodities",),
            "foreign_exchange": ("foreign_exchange",),
            "policy_rate": ("policy_rates",),
            "money_market": ("money_market",),
        }
        rows = []
        precision_sentinel = 0.0123456789012345
        for asset_class, count in class_counts.items():
            asset_groups = groups[asset_class]
            for index in range(count):
                rows.append({
                    "asset_class": asset_class,
                    "group": asset_groups[index % len(asset_groups)],
                    "series_code": f"{asset_class}_{index}",
                    "name_cn": f"指标{len(rows)}",
                    "level_unit": "percent",
                    "change_unit": "pct" if asset_class in {"commodity", "foreign_exchange"} else "bp",
                    "sort_order": index + 1,
                    "daily_change": float(index),
                    "weekly_change": float(index),
                    "mtd_change": float(index),
                    "ytd_change": float(index),
                    "qc_flag": "FETCH_FAILED" if index == 0 else "OK",
                })
        rows[1]["daily_change"] = precision_sentinel
        detail = pd.DataFrame(rows)
        source_log = detail[["asset_class", "series_code"]].copy()
        source_log["status"] = detail["qc_flag"].map(
            {"OK": "OK", "FETCH_FAILED": "FETCH_FAILED"}
        )

        with tempfile.TemporaryDirectory() as directory, patch(
            "pipeline.internal.scripts.fetch_macro_assets.fetch_macro_assets",
            return_value=(detail, source_log),
        ), patch.object(
            os.sys, "argv", ["fetch_macro_assets.py", "--output-dir", directory]
        ), patch("sys.stdout", new_callable=StringIO) as stdout:
            fetch_cli.main()
            root = Path(directory)
            snapshot = json.loads(
                (root / "macro_assets_snapshot.json").read_text(encoding="utf-8"),
                parse_constant=lambda value: (_ for _ in ()).throw(
                    ValueError(f"non-finite constant: {value}")
                ),
            )

            self.assertEqual(len(pd.read_csv(root / "fixed_income.csv")), 20)
            self.assertEqual(len(pd.read_csv(root / "commodities.csv")), 15)
            self.assertEqual(len(pd.read_csv(root / "foreign_exchange.csv")), 4)
            self.assertEqual(len(pd.read_csv(root / "policy_rates.csv")), 12)
            self.assertEqual(len(pd.read_csv(root / "money_market.csv")), 7)
            self.assertEqual(len(pd.read_csv(root / "source_log.csv")), 58)
            self.assertEqual(
                sum(len(snapshot[key]) for key in (
                    "fixed_income", "commodities", "foreign_exchange", "policy_rates", "money_market"
                )),
                58,
            )
            self.assertEqual(len(snapshot["macro_divergence"]), 28)
            self.assertEqual(len(snapshot["source_log"]), 58)
            self.assertLess(
                abs(snapshot["fixed_income"][1]["daily_change"] - precision_sentinel),
                1e-12,
            )
            output = stdout.getvalue()
            for asset_class, configured in class_counts.items():
                self.assertIn(
                    f"{asset_class}: configured={configured}, fetched={configured - 1}, failed=1",
                    output,
                )

    def test_cli_rolls_back_whole_output_bundle_when_publish_fails(self):
        from pipeline.internal.scripts import fetch_macro_assets as fetch_cli

        detail = pd.DataFrame({
            "asset_class": ["fixed_income"], "group": ["sovereign_curve"],
            "series_code": ["UST10Y"], "name_cn": ["美国10年期国债收益率"],
            "level_unit": ["percent"], "change_unit": ["bp"], "sort_order": [1],
            "daily_change": [1.0], "weekly_change": [2.0],
            "mtd_change": [3.0], "ytd_change": [4.0], "qc_flag": ["OK"],
        })
        source_log = pd.DataFrame({"series_code": ["UST10Y"], "status": ["OK"]})
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory) / "published"
            (root / "raw").mkdir(parents=True)
            (root / "marker.bin").write_bytes(b"old bundle\x00")
            (root / "raw" / "old.raw").write_bytes(b"old raw\xff")
            before = {
                path.relative_to(root): path.read_bytes()
                for path in root.rglob("*") if path.is_file()
            }
            real_replace = fetch_cli.os.replace

            def fail_staging_publish(source, target):
                if Path(target) == root and ".staging-" in Path(source).name:
                    raise OSError("publish failed")
                return real_replace(source, target)

            with patch(
                "pipeline.internal.scripts.fetch_macro_assets.fetch_macro_assets",
                return_value=(detail, source_log),
            ), patch(
                "pipeline.internal.scripts.fetch_macro_assets.os.replace", side_effect=fail_staging_publish
            ), patch.object(
                os.sys, "argv", ["fetch_macro_assets.py", "--output-dir", str(root)]
            ):
                with self.assertRaisesRegex(OSError, "publish failed"):
                    fetch_cli.main()

            after = {
                path.relative_to(root): path.read_bytes()
                for path in root.rglob("*") if path.is_file()
            }
            self.assertEqual(after, before)

    def test_cli_writes_public_tables_and_official_commodity_history(self):
        from pipeline.internal.scripts import fetch_macro_assets as fetch_cli

        detail = pd.DataFrame(
            [
                {
                    "asset_class": "liquidity",
                    "group": "fed_liquidity",
                    "series_code": "FED_TOTAL_ASSETS",
                    "name_cn": "美联储总资产",
                    "level_unit": "usd_billions",
                    "change_unit": "usd_billions",
                    "sort_order": 1,
                    "daily_change": 10.0,
                    "weekly_change": 20.0,
                    "mtd_change": 30.0,
                    "ytd_change": 40.0,
                    "qc_flag": "OK",
                },
                {
                    "asset_class": "cross_asset",
                    "group": "cross_asset_correlation",
                    "series_code": "US_STOCK_BOND_CORR_13W",
                    "name_cn": "股债相关性",
                    "level_unit": "correlation",
                    "change_unit": "correlation_points",
                    "sort_order": 2,
                    "daily_change": 0.01,
                    "weekly_change": 0.02,
                    "mtd_change": 0.03,
                    "ytd_change": 0.04,
                    "qc_flag": "OK",
                },
                {
                    "asset_class": "commodity",
                    "group": "commodities",
                    "series_code": "COMEX_COPPER",
                    "name_cn": "铜代理",
                    "level_unit": "usd_per_pound",
                    "change_unit": "pct",
                    "sort_order": 3,
                    "daily_change": float("nan"),
                    "weekly_change": float("nan"),
                    "mtd_change": float("nan"),
                    "ytd_change": float("nan"),
                    "qc_flag": "FETCH_FAILED",
                },
                {
                    "asset_class": "calculation_input",
                    "group": "correlation_inputs",
                    "series_code": "SPY_CLOSE_PROXY",
                    "name_cn": "隐藏输入",
                    "level_unit": "usd_per_share",
                    "change_unit": "pct",
                    "sort_order": 4,
                    "daily_change": 0.01,
                    "weekly_change": 0.02,
                    "mtd_change": 0.03,
                    "ytd_change": 0.04,
                    "qc_flag": "OK",
                },
            ]
        )
        source_log = pd.DataFrame(
            {
                "series_code": detail["series_code"],
                "status": ["OK", "OK", "FETCH_FAILED", "OK"],
                "requiredness": ["required", "optional", "optional", "optional"],
            }
        )
        commodity_price_history = pd.DataFrame(
            [
                {
                    "record_id": "price:WTI:2026-08-28",
                    "as_of_date": "2026-08-30",
                    "commodity_code": "WTI",
                    "commodity_family": "refined_products",
                    "series_code": "WTI",
                    "price_kind": "official_cash",
                    "observation_date": "2026-08-28",
                    "known_as_of": "2026-08-28T18:00:00Z",
                    "value": 64.75,
                    "unit": "$/BBL",
                    "source": "U.S. Energy Information Administration",
                    "source_url": "https://api.eia.gov/v2/petroleum/pri/spt/data/",
                    "qc_flag": "OK",
                }
            ],
            columns=macro_assets_module.PRICE_HISTORY_FIELDS,
        )
        bundle = MacroAssetBundle(
            detail=detail,
            source_log=source_log,
            commodity_price_history=commodity_price_history,
        )

        with tempfile.TemporaryDirectory() as directory, patch(
            "pipeline.internal.scripts.fetch_macro_assets.fetch_macro_assets",
            return_value=bundle,
        ), patch.object(
            os.sys, "argv", ["fetch_macro_assets.py", "--output-dir", directory]
        ):
            fetch_cli.main()
            root = Path(directory)
            self.assertTrue((root / "liquidity.csv").exists())
            self.assertTrue((root / "cross_asset.csv").exists())
            self.assertEqual(
                pd.read_csv(root / "liquidity.csv")["series_code"].tolist(),
                ["FED_TOTAL_ASSETS"],
            )
            self.assertEqual(
                pd.read_csv(root / "cross_asset.csv")["series_code"].tolist(),
                ["US_STOCK_BOND_CORR_13W"],
            )
            self.assertTrue(pd.read_csv(root / "commodities.csv").empty)
            self.assertEqual(
                pd.read_csv(root / "commodity_price_history.csv")[
                    "record_id"
                ].tolist(),
                ["price:WTI:2026-08-28"],
            )
            snapshot = json.loads(
                (root / "macro_assets_snapshot.json").read_text(encoding="utf-8")
            )

        self.assertEqual(len(snapshot["liquidity"]), 1)
        self.assertEqual(len(snapshot["cross_asset"]), 1)
        self.assertEqual(snapshot["commodities"], [])
        self.assertEqual(
            snapshot["commodity_price_history"][0]["record_id"],
            "price:WTI:2026-08-28",
        )
        self.assertEqual(len(snapshot["source_log"]), 4)


class MacroAssetHistoryBundleTests(unittest.TestCase):
    @staticmethod
    def _config(
        series_code: str,
        provider: str,
        *,
        commodity_code: str,
        commodity_family: str,
        frequency: str,
        price_kind: str,
        level_unit: str,
        sort_order: int,
    ) -> MacroAssetConfig:
        return MacroAssetConfig(
            asset_class="commodity",
            group="commodities",
            series_code=series_code,
            name_cn=series_code,
            name_en=series_code,
            provider=provider,
            provider_symbol=series_code,
            source="Official fixture" if provider != "yahoo_chart" else "Vendor fixture",
            source_url="https://official.example.test/data",
            frequency=frequency,
            level_unit=level_unit,
            change_unit="pct",
            sort_order=sort_order,
            commodity_code=commodity_code,
            commodity_family=commodity_family,
            price_kind=price_kind,
            freshness_days="45" if provider == "world_bank_pink_sheet" else "",
        )

    def test_bundle_publishes_only_configured_eia_and_world_bank_histories(self):
        as_of = date(2026, 8, 30)
        universe = [
            self._config(
                "WTI",
                "eia_v2",
                commodity_code="WTI",
                commodity_family="refined_products",
                frequency="daily",
                price_kind="official_cash",
                level_unit="$/BBL",
                sort_order=1,
            ),
            self._config(
                "COMEX_GOLD",
                "world_bank_pink_sheet",
                commodity_code="GOLD_COMEX",
                commodity_family="gold",
                frequency="monthly",
                price_kind="official_monthly_benchmark",
                level_unit="$/troy oz",
                sort_order=2,
            ),
            self._config(
                "BTC_USD",
                "yahoo_chart",
                commodity_code="BTC_USD",
                commodity_family="digital_asset",
                frequency="daily",
                price_kind="",
                level_unit="usd",
                sort_order=3,
            ),
        ]

        def history(config, _session, as_of_date=None):
            del as_of_date
            if config.frequency == "monthly":
                points = [
                    {"date": date(2026, 7, 31), "value": 3300.0, "unit": config.level_unit},
                    {"date": date(2026, 8, 29), "value": 3400.0, "unit": config.level_unit},
                ]
            else:
                points = [
                    {"date": date(2026, 8, 22), "value": 70.0, "unit": config.level_unit},
                    {"date": date(2026, 8, 29), "value": 75.0, "unit": config.level_unit},
                ]
            return points, b"official raw", f"https://official.example.test/{config.series_code}"

        with patch.object(
            macro_assets_module,
            "load_macro_asset_universe",
            return_value=universe,
        ), patch.object(
            macro_assets_module,
            "_fetch_config_history",
            side_effect=history,
        ):
            bundle = fetch_macro_asset_bundle(as_of_date=as_of)
            legacy = fetch_macro_assets(as_of_date=as_of)

        self.assertIsInstance(bundle, MacroAssetBundle)
        self.assertEqual(
            set(bundle.commodity_price_history["commodity_code"]),
            {"WTI", "GOLD_COMEX"},
        )
        self.assertEqual(
            set(bundle.commodity_price_history["series_code"]),
            {"WTI", "COMEX_GOLD"},
        )
        self.assertNotIn("BTC_USD", set(bundle.commodity_price_history["commodity_code"]))
        self.assertIs(type(legacy), tuple)
        self.assertEqual(len(legacy), 2)
        self.assertIsInstance(legacy[0], pd.DataFrame)
        self.assertIsInstance(legacy[1], pd.DataFrame)

    def test_future_known_as_of_cannot_become_the_macro_snapshot(self):
        as_of = date(2026, 8, 30)
        universe = [
            self._config(
                "WTI",
                "eia_v2",
                commodity_code="WTI",
                commodity_family="refined_products",
                frequency="daily",
                price_kind="official_cash",
                level_unit="$/BBL",
                sort_order=1,
            )
        ]
        for invalid_known_as_of in (
            "2026-08-31T00:00:00Z",
            "2026-08-29T12:00:00",
        ):
            with self.subTest(known_as_of=invalid_known_as_of):
                history = [
                    {
                        "date": date(2026, 8, 22),
                        "known_as_of": "2026-08-22T12:00:00Z",
                        "value": 70.0,
                        "unit": "$/BBL",
                    },
                    {
                        "date": date(2026, 8, 29),
                        "known_as_of": invalid_known_as_of,
                        "value": 99.0,
                        "unit": "$/BBL",
                    },
                ]

                with patch.object(
                    macro_assets_module,
                    "load_macro_asset_universe",
                    return_value=universe,
                ), patch.object(
                    macro_assets_module,
                    "_fetch_config_history",
                    return_value=(
                        history,
                        b"official raw",
                        "https://official.example.test/WTI",
                    ),
                ):
                    bundle = fetch_macro_asset_bundle(
                        as_of_date=as_of,
                        allow_partial=True,
                    )

                self.assertEqual(
                    bundle.detail.loc[0, "qc_flag"],
                    "FETCH_FAILED",
                )
                self.assertIsNone(bundle.detail.loc[0, "latest_value"])
                self.assertIn(
                    "known_as_of",
                    bundle.source_log.loc[0, "error"],
                )
                self.assertEqual(
                    bundle.source_log.loc[0, "status"],
                    "FETCH_FAILED",
                )
                self.assertTrue(bundle.commodity_price_history.empty)

    def test_config_known_as_of_is_validated_before_snapshot_fallback(self):
        as_of = date(2026, 8, 30)
        base_config = self._config(
            "WTI",
            "eia_v2",
            commodity_code="WTI",
            commodity_family="refined_products",
            frequency="daily",
            price_kind="official_cash",
            level_unit="$/BBL",
            sort_order=1,
        )
        history = [
            {
                "date": date(2026, 8, 29),
                "value": 99.0,
                "unit": "$/BBL",
            }
        ]

        for invalid_known_as_of in (
            "2026-08-31T00:00:00Z",
            "2026-08-29T12:00:00",
        ):
            with self.subTest(known_as_of=invalid_known_as_of):
                config = replace(
                    base_config,
                    known_as_of=invalid_known_as_of,
                )
                with patch.object(
                    macro_assets_module,
                    "load_macro_asset_universe",
                    return_value=[config],
                ), patch.object(
                    macro_assets_module,
                    "_fetch_config_history",
                    return_value=(
                        history,
                        b"official raw",
                        "https://official.example.test/WTI",
                    ),
                ), patch.object(
                    macro_assets_module,
                    "calculate_macro_snapshot",
                    wraps=macro_assets_module.calculate_macro_snapshot,
                ) as snapshot:
                    bundle = fetch_macro_asset_bundle(
                        as_of_date=as_of,
                        allow_partial=True,
                    )

                snapshot.assert_not_called()
                self.assertEqual(
                    bundle.detail.loc[0, "qc_flag"],
                    "FETCH_FAILED",
                )
                self.assertIsNone(bundle.detail.loc[0, "latest_value"])

    def test_future_observation_is_cut_off_before_snapshot(self):
        as_of = date(2026, 8, 30)
        config = self._config(
            "WTI",
            "eia_v2",
            commodity_code="WTI",
            commodity_family="refined_products",
            frequency="daily",
            price_kind="official_cash",
            level_unit="$/BBL",
            sort_order=1,
        )
        history = [
            {
                "date": date(2026, 8, 22),
                "value": 65.0,
                "unit": "$/BBL",
            },
            {
                "date": date(2026, 8, 29),
                "value": 70.0,
                "unit": "$/BBL",
            },
            {
                "date": date(2026, 8, 31),
                "value": 99.0,
                "unit": "$/BBL",
            },
        ]

        with patch.object(
            macro_assets_module,
            "load_macro_asset_universe",
            return_value=[config],
        ), patch.object(
            macro_assets_module,
            "_fetch_config_history",
            return_value=(
                history,
                b"official raw",
                "https://official.example.test/WTI",
            ),
        ):
            try:
                bundle = fetch_macro_asset_bundle(as_of_date=as_of)
            except ValueError as error:
                self.fail(f"future observation blocked publication: {error}")

        self.assertNotEqual(bundle.detail.loc[0, "qc_flag"], "FETCH_FAILED")
        self.assertEqual(bundle.detail.loc[0, "latest_date"], "2026-08-29")
        self.assertEqual(bundle.detail.loc[0, "latest_value"], 70.0)
        self.assertEqual(
            bundle.commodity_price_history["observation_date"].tolist(),
            ["2026-08-22", "2026-08-29"],
        )


if __name__ == "__main__":
    unittest.main()
